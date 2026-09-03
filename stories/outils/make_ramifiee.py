#!/usr/bin/env python3
"""Construit un arbre ramifié 3×3×3 (27 feuilles) autour d'une leçon AcoMytha."""
from __future__ import annotations

import argparse
import json
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]


def _load_lessons():
    xlsx = ROOT / "referentiel" / "lecons.xlsx"
    if xlsx.exists():
        from openpyxl import load_workbook

        wb = load_workbook(xlsx, read_only=True, data_only=True)
        rows = list(wb["lecons"].iter_rows(values_only=True))
        headers = [str(h) for h in rows[0]]
        out = {}
        for r in rows[1:]:
            d = {headers[i]: r[i] for i in range(len(headers))}
            lid = d.get("lesson_id")
            if not lid:
                continue
            for k in (
                "required_messages",
                "safe_actions",
                "misconceptions",
                "forbidden_in_audio",
                "answer_intents",
                "compatible_lessons",
            ):
                v = d.get(k) or ""
                d[k] = [x.strip() for x in str(v).split("|") if x.strip()]
            out[str(lid)] = d
        wb.close()
        return out
    data = json.loads((ROOT / "referentiel" / "lecons.json").read_text(encoding="utf-8"))
    return {l["lesson_id"]: l for l in data["lessons"]}


BY_ID = _load_lessons()

# Choix narratifs neutres (jamais un test de sécurité)
CHOICE_SETS = {
    "lieux_maison": ("la cuisine", "le jardin", "la chambre"),
    "lieux_parc": ("le bac à sable", "le toboggan", "les balançoires"),
    "objets": ("le ballon", "le seau", "le doudou"),
    "jeux": ("les cubes", "le livre", "la dînette"),
    "goûter": ("une pomme", "un yaourt", "un morceau de pain"),
    "camarades": ("Tom", "Léa", "Sami"),
    "moments": ("le matin", "après la sieste", "le soir"),
    "animaux": ("le chat", "le chien", "la poule"),
    "couleurs": ("rouge", "bleu", "vert"),
    "transports": ("le train", "le bus", "la voiture"),
}

CHARACTERS = [
    ("Lina", "elle"), ("Nora", "elle"), ("Maya", "elle"), ("Inès", "elle"),
    ("Léa", "elle"), ("Zoé", "elle"), ("Sara", "elle"), ("Lila", "elle"),
    ("Tom", "il"), ("Sami", "il"), ("Hugo", "il"), ("Jules", "il"),
    ("Noé", "il"), ("Adam", "il"), ("Nino", "il"), ("Kenzo", "il"),
]

SENS = [
    "On entend un oiseau.", "Ça sent le pain chaud.", "Le vent est doux.",
    "Le sol est un peu froid.", "Une feuille tombe.", "L'eau coule, tout petit bruit.",
    "Un vélo passe au loin.", "La lumière est claire.", "Il y a des miettes sur la table.",
    "Un chat miaule une fois.", "Les chaussures font toc toc.", "La couverture est douce.",
]

BEATS = [
    "On rit un peu.", "On se tait une seconde.", "On se tient la main.",
    "On range un objet.", "On dit merci.", "On souffle doucement.",
    "On écoute jusqu'au bout.", "On attend son tour.", "On montre du doigt, sans courir.",
]


def join_messages(lesson: dict) -> str:
    msgs = lesson.get("required_messages") or []
    bits = [m.replace(";", " ").strip() for m in msgs]
    return ", ".join(bits)


def safe_line(lesson: dict) -> str:
    acts = [a.strip() for a in (lesson.get("safe_actions") or []) if a.strip()]
    if not acts:
        return "On fait le geste sûr, avec papa ou maman."
    # actions already contain verbs or noun phrases — do not prefix "On" blindly
    spoken = []
    for a in acts[:3]:
        if a[0].isupper() or a.startswith(("on ", "On ", "s'", "se ", "dire", "aller", "attendre", "prendre")):
            spoken.append(a)
        else:
            spoken.append(a)
    return "Voici le geste : " + " ; ".join(spoken) + "."


def question_for(lesson: dict) -> dict:
    intents = lesson.get("answer_intents") or ["OUI"]
    title = lesson.get("title_child_audio") or lesson.get("title")
    prompt = f"Que fait-on ? {title}."
    # shorter prompts by domain
    examples = []
    for it in intents[:3]:
        examples.append(it.lower().replace("_", " "))
    feedback = safe_line(lesson)
    return {
        "prompt": prompt if len(prompt) < 80 else f"Que fait-on maintenant ?",
        "expected_intents": intents,
        "accepted_examples": examples or ["oui"],
        "retry_prompt": "Dis-le avec un petit mot.",
        "positive_feedback": f"Oui. {feedback}",
        "wrong_feedback": feedback,
    }


def build_tree(
    tree_id: str,
    lesson_id: str,
    age_band: str,
    setting: str,
    hero: str,
    pronoun: str,
    choice_keys: tuple[str, str, str],
    secondary: str | None = None,
    index: int = 1,
) -> dict:
    lesson = BY_ID[lesson_id]
    q = question_for(lesson)
    req = join_messages(lesson)
    safe = safe_line(lesson)
    adult = "maman" if index % 2 else "papa"
    other = "papa" if adult == "maman" else "maman"
    title = lesson.get("title_child_audio") or lesson["title"]
    c1 = CHOICE_SETS[choice_keys[0]]
    c2 = CHOICE_SETS[choice_keys[1]]
    c3 = CHOICE_SETS[choice_keys[2]]

    # pronouns
    va = "va" if pronoun == "elle" or True else "va"
    son = "sa" if pronoun == "elle" else "son"
    the = hero

    nodes = {}

    def audio(nid, text, nxt):
        nodes[nid] = {"id": nid, "type": "audio", "text": text, "next": nxt}

    def ending(nid, text):
        nodes[nid] = {"id": nid, "type": "ending", "text": text}

    def choice(nid, prompt, labels, nexts, default):
        opts = []
        for i, (lab, nxt) in enumerate(zip(labels, nexts)):
            intent = lab.upper().replace(" ", "_").replace("'", "")[:24]
            opts.append({"id": f"{nid}_o{i+1}", "label": lab, "intent": intent, "next": nxt})
        nodes[nid] = {
            "id": nid,
            "type": "choice_story",
            "prompt": prompt,
            "options": opts,
            "default_next": default,
        }

    def qn(nid, nxt):
        nodes[nid] = {
            "id": nid,
            "type": "question_lesson",
            "prompt": q["prompt"],
            "expected_intents": q["expected_intents"],
            "accepted_examples": q["accepted_examples"],
            "retry_prompt": q["retry_prompt"],
            "positive_feedback": q["positive_feedback"],
            "wrong_feedback": q["wrong_feedback"],
            "default_next": nxt,
            "next": nxt,
            "lesson_id": lesson_id,
        }

    def fb(nid, text, nxt):
        nodes[nid] = {"id": nid, "type": "feedback", "text": text, "next": nxt}

    # IDs never collide: ch1 / brA / qA / fbA / ch2A / brA1 / ch3A1 / endA1X
    audio(
        "root",
        f"Aujourd'hui, {the} est avec {adult}, {setting}. "
        f"{the} a envie de jouer. "
        f"{adult} dit : Je suis là, tout près. On joue ensemble. "
        f"On va apprendre : {title}. {safe} "
        f"{the} écoute {adult}. {the} entend les mots : {req}.",
        "ch1",
    )
    choice(
        "ch1",
        f"{the} choisit. {c1[0]}, {c1[1]}, ou {c1[2]} ?",
        list(c1),
        ["brA", "brB", "brC"],
        "brA",
    )

    letters = ["A", "B", "C"]
    for i, L in enumerate(letters):
        sens = SENS[(index + i) % len(SENS)]
        beat = BEATS[(index + i) % len(BEATS)]
        audio(
            f"br{L}",
            f"{the} va vers {c1[i]}. {sens} "
            f"{adult} dit : Bravo. Tu es arrivé. Tu as fait du bon travail. "
            f"Ici, c'est {c1[i]}. Ce n'est pas comme les autres endroits. "
            f"{the} se souvient : {req}. {safe} "
            f"{beat} {adult} dit : Je suis là. On reste ensemble.",
            f"q{L}",
        )
        qn(f"q{L}", f"fb{L}")
        fb(
            f"fb{L}",
            f"Oui. {safe} {the} respire. {the} retient : {req}. On continue, avec {adult}.",
            f"ch2{L}",
        )
        choice(
            f"ch2{L}",
            f"Ensuite, {the} choisit {c2[0]}, {c2[1]}, ou {c2[2]} ?",
            list(c2),
            [f"br{L}1", f"br{L}2", f"br{L}3"],
            f"br{L}1",
        )
        for j in range(3):
            nid = f"{L}{j+1}"
            sens2 = SENS[(index + i * 3 + j + 2) % len(SENS)]
            beat2 = BEATS[(index + j + 3) % len(BEATS)]
            audio(
                f"br{nid}",
                f"{the} a choisi {c2[j]}. {sens2} "
                f"{c2[j][0].upper() + c2[j][1:]} reste à sa place, près de {adult}. "
                f"{adult} dit : Tu as fini ? On le fait ensemble. "
                f"{safe} "
                f"{the} répète tout bas : {req}. {beat2} "
                f"{other} n'est pas loin.",
                f"ch3{nid}",
            )
            choice(
                f"ch3{nid}",
                f"Pour finir, {the} va vers {c3[0]}, {c3[1]}, ou {c3[2]} ?",
                list(c3),
                [f"end{nid}X", f"end{nid}Y", f"end{nid}Z"],
                f"end{nid}X",
            )
            for k, suf in enumerate(("X", "Y", "Z")):
                sens3 = SENS[(index + i + j + k + 5) % len(SENS)]
                beat3 = BEATS[(index + k + i) % len(BEATS)]
                ending(
                    f"end{nid}{suf}",
                    f"{the} rejoint {c3[k]}. {sens3} {beat3} "
                    f"Cette fin-là est celle de {c1[i]}, {c2[j]} et {c3[k]}. "
                    f"{the} a appris : {title}. {safe} "
                    f"{the} a entendu : {req}. "
                    f"{the} dit merci à {adult}. On rentre. L'histoire est finie.",
                )

    secondary_lessons = [secondary] if secondary else []
    return {
        "tree_id": tree_id,
        "kind": "ramifiee",
        "version": 1,
        "language": "fr",
        "age_band": age_band,
        "age_range": {"N1": "3-4", "N2": "4-5", "N3": "5-6"}[age_band],
        "title": f"{title} — {setting}",
        "lesson_id": lesson_id,
        "secondary_lessons": secondary_lessons,
        "domain": lesson["domain_id"],
        "subdomain": lesson["subdomain_id"],
        "framing": lesson.get("framing") or "standard",
        "sensitivity": lesson.get("sensitivity") or "standard",
        "family_model": "father_mother_children",
        "setting": setting,
        "characters": [hero, adult, other],
        "root_id": "root",
        "duration_estimate_seconds": {"min": 180, "avg": 300, "max": 480},
        "nodes": nodes,
        "validation": {"status": "PENDING", "blocking_findings": 0},
    }


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--lesson", required=True)
    p.add_argument("--tree-id", required=True)
    p.add_argument("--age", default="N2")
    p.add_argument("--setting", default="à la maison")
    p.add_argument("--hero", default="Lina")
    p.add_argument("--pronoun", default="elle")
    p.add_argument("--choices", default="lieux_maison,objets,camarades")
    p.add_argument("--secondary", default="")
    p.add_argument("--index", type=int, default=1)
    p.add_argument("--out", required=True)
    args = p.parse_args()
    keys = tuple(args.choices.split(","))
    tree = build_tree(
        args.tree_id, args.lesson, args.age, args.setting,
        args.hero, args.pronoun, keys, args.secondary or None, args.index,
    )
    out = Path(args.out)
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(tree, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(out, "nodes", len(tree["nodes"]))


if __name__ == "__main__":
    main()
