#!/usr/bin/env python3
"""F-NAR-008 — dump / merge agents / apply (l'xlsx source reste intact jusqu'à apply)."""

from __future__ import annotations

import argparse
import json
import shutil
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
ARBRES = ROOT / "arbres"
ARCHIVE = ROOT / "archive" / "arbres"
REWRITES = ROOT / "rewrites"


def xlsx_path(story_id: str) -> Path:
    live = ARBRES / f"{story_id}.xlsx"
    if live.exists():
        return live
    archived = ARCHIVE / f"{story_id}.xlsx"
    if archived.exists():
        return archived
    raise SystemExit(f"absent: {story_id}.xlsx (arbres/ et archive/)")

CHUNK_FIELDS = (
    "chunk_id",
    "kind",
    "text",
    "script",
    "sons",
    "length_scale_piper",
    "rate_label",
    "pause_after_ms",
    "option_1_label",
    "option_2_label",
    "option_3_label",
    "expected_answer",
    "accepted_examples",
    "retry_prompt",
    "text_ssml",
    "text_xai_tags",
    "rate_wpm",
    "speed_xai",
    "pitch_label",
    "pitch_ssml",
    "pitch_xai_tag",
    "volume_label",
    "volume_db",
    "emphasis_words",
    "pause_before_ms",
    "pause_sentence_ms",
    "style_energy",
    "style_contour",
    "noise_scale_piper",
    "kokoro_speed",
    "melo_speed",
    "espeak_amp",
    "espeak_pitch",
    "espeak_word_gap",
    "notes",
    "option_1_next_chunk",
    "option_2_next_chunk",
    "option_3_next_chunk",
    "default_next_chunk",
)


def dump_story(story_id: str) -> Path:
    src = xlsx_path(story_id)
    wb = load_workbook(src, read_only=True, data_only=True)
    meta = {}
    if "meta" in wb.sheetnames:
        for row in wb["meta"].iter_rows(values_only=True):
            if row and row[0] not in (None, "clé"):
                meta[str(row[0])] = "" if row[1] is None else str(row[1])
    rows = list(wb["chunks"].iter_rows(values_only=True))
    headers = [str(h) if h else "" for h in rows[0]]
    chunks = []
    for r in rows[1:]:
        d = {headers[i]: r[i] if i < len(r) else None for i in range(len(headers))}
        if not d.get("chunk_id"):
            continue
        chunks.append({k: d.get(k) for k in CHUNK_FIELDS if k in d})
    wb.close()
    out_dir = REWRITES / story_id
    out_dir.mkdir(parents=True, exist_ok=True)
    payload = {
        "story_id": story_id,
        "fil_rouge": "",
        "title": meta.get("title", ""),
        "lesson_id": meta.get("lesson_id", ""),
        "age_band": meta.get("age_band", ""),
        "kind": meta.get("kind", ""),
        "characters": meta.get("characters", ""),
        "setting": meta.get("setting", ""),
        "chunks": chunks,
    }
    path = out_dir / "source.json"
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return path


def _score(agent: dict) -> int:
    """Plus le fil rouge et les textes sont longs (sans être vides), mieux c'est."""
    s = len(agent.get("fil_rouge") or "")
    for c in agent.get("chunks") or []:
        s += min(len(str(c.get("text") or "")), 400)
        s += 20 if c.get("script") else 0
    return s


def merge_story(story_id: str) -> Path:
    folder = REWRITES / story_id
    source = json.loads((folder / "source.json").read_text(encoding="utf-8"))
    agents = []
    for p in sorted(folder.glob("agent_*.json")):
        agents.append(json.loads(p.read_text(encoding="utf-8")))
    if not agents:
        raise SystemExit(f"aucun agent_*.json dans {folder}")
    agents.sort(key=_score, reverse=True)
    best = agents[0]
    by_id = {c["chunk_id"]: c for c in source["chunks"]}
    for ag in reversed(agents):
        for c in ag.get("chunks") or []:
            cid = c.get("chunk_id")
            if cid in by_id:
                for k in CHUNK_FIELDS:
                    if c.get(k) not in (None, ""):
                        by_id[cid][k] = c[k]
    merged = dict(source)
    merged["fil_rouge"] = best.get("fil_rouge") or source.get("fil_rouge")
    merged["title"] = best.get("title") or source.get("title")
    merged["merged_from"] = [p.name for p in sorted(folder.glob("agent_*.json"))]
    merged["chunks"] = [by_id[c["chunk_id"]] for c in source["chunks"]]
    out = folder / "merged.json"
    out.write_text(json.dumps(merged, ensure_ascii=False, indent=2), encoding="utf-8")
    return out


def apply_story(story_id: str) -> Path:
    folder = REWRITES / story_id
    merged_path = folder / "merged.json"
    if not merged_path.exists():
        merge_story(story_id)
    merged = json.loads(merged_path.read_text(encoding="utf-8"))
    src = xlsx_path(story_id)
    backup = folder / "original.xlsx"
    if not backup.exists():
        shutil.copy2(src, backup)
    wb = load_workbook(src)
    if "meta" in wb.sheetnames:
        ms = wb["meta"]
        meta_map = {
            "title": merged.get("title") or "",
            "fil_rouge": merged.get("fil_rouge") or "",
            "characters": merged.get("characters") or "",
            "setting": merged.get("setting") or "",
        }
        if merged.get("secondary_lessons") is not None:
            meta_map["secondary_lessons"] = merged.get("secondary_lessons") or ""
        seen: set[str] = set()
        for row in ms.iter_rows(min_col=1, max_col=2):
            key = row[0].value
            if key in meta_map:
                row[1].value = meta_map[key]
                seen.add(str(key))
        for key, val in meta_map.items():
            if key not in seen and val:
                ms.append([key, val])
    ws = wb["chunks"]
    headers = [c.value for c in ws[1]]
    for name in ("script", "sons", "length_scale_piper", "rate_label"):
        if name not in headers:
            ws.cell(1, len(headers) + 1, name)
            headers.append(name)
    idx = {name: i + 1 for i, name in enumerate(headers)}
    by_id = {c["chunk_id"]: c for c in merged["chunks"]}
    for r in range(2, ws.max_row + 1):
        cid = ws.cell(r, idx["chunk_id"]).value
        if cid not in by_id:
            continue
        c = by_id[cid]
        for k in (
            "text",
            "script",
            "sons",
            "length_scale_piper",
            "rate_label",
            "pause_after_ms",
            "option_1_label",
            "option_2_label",
            "option_3_label",
            "expected_answer",
            "accepted_examples",
            "retry_prompt",
            "text_ssml",
            "text_xai_tags",
            "rate_wpm",
            "speed_xai",
            "pitch_label",
            "pitch_ssml",
            "pitch_xai_tag",
            "volume_label",
            "volume_db",
            "emphasis_words",
            "pause_before_ms",
            "pause_sentence_ms",
            "style_energy",
            "style_contour",
            "noise_scale_piper",
            "kokoro_speed",
            "melo_speed",
            "espeak_amp",
            "espeak_pitch",
            "espeak_word_gap",
            "notes",
        ):
            if k in idx and c.get(k) not in (None, ""):
                ws.cell(r, idx[k], c[k])
                if k == "text" and "text_ssml" in idx and not c.get("text_ssml"):
                    ws.cell(r, idx["text_ssml"], c[k])
    if "journal" in wb.sheetnames:
        wb["journal"].append(["F-NAR-008 récit, pas une leçon en puces"])
    wb.save(src)
    wb.close()
    dest = ARBRES / f"{story_id}.xlsx"
    if src.resolve() != dest.resolve():
        dest.parent.mkdir(parents=True, exist_ok=True)
        shutil.move(str(src), str(dest))
        src = dest
    return src


def dump_list(path: Path) -> int:
    ids = [ln.strip() for ln in path.read_text(encoding="utf-8").splitlines() if ln.strip()]
    n = 0
    for sid in ids:
        dump_story(sid)
        n += 1
    return n


def apply_ready() -> list[str]:
    done = []
    for folder in sorted(REWRITES.iterdir()):
        if not folder.is_dir():
            continue
        merged = folder / "merged.json"
        if not merged.exists():
            continue
        sid = folder.name
        try:
            xlsx_path(sid)
        except SystemExit:
            continue
        apply_story(sid)
        done.append(sid)
    return done


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("cmd", choices=["dump", "merge", "apply", "dump-list", "apply-ready"])
    ap.add_argument("story_id", nargs="?")
    args = ap.parse_args()
    if args.cmd == "dump-list":
        if not args.story_id:
            raise SystemExit("dump-list <fichier d'ids>")
        print(dump_list(Path(args.story_id)))
    elif args.cmd == "apply-ready":
        done = apply_ready()
        print("\n".join(done))
        print(f"applied {len(done)}")
    elif args.cmd == "dump":
        print(dump_story(args.story_id))
    elif args.cmd == "merge":
        print(merge_story(args.story_id))
    else:
        print(apply_story(args.story_id))


if __name__ == "__main__":
    main()
