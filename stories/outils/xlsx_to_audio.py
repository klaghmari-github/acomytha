#!/usr/bin/env python3
"""Bake audio depuis les xlsx.

Piper sort du 22050 Hz (souvent muet sur téléphone / Windows / navigateur).
On réécrit en WAV PCM 44100 Hz 16-bit + MP3 64 kbit/s, niveau normalisé.
"""
from __future__ import annotations

import argparse
import shutil
import struct
import subprocess
import sys
import tempfile
import wave
from dataclasses import dataclass
from pathlib import Path

import numpy as np
from openpyxl import load_workbook
from scipy.signal import resample, resample_poly

sys.path.insert(0, str(Path(__file__).resolve().parent))
from voice_cast import parse_roster, script_load, split_script

ROOT = Path(__file__).resolve().parents[1]
ARBRES = ROOT / "arbres"
AUDIO = ROOT / "audio"
VOICES = ROOT / "outils" / "voices"
TARGET_SR = 44100
PEAK_TARGET = 0.89  # ~ -1 dBFS
MIN_PEAK = 500  # en dessous = considéré muet
# Mono = durée qui défile, silence sur iPhone / Windows / BT / HDMI.
CHANNELS = 2
GAP_S = 0.28


@dataclass
class Voice:
    model: str
    speaker: int | None = None
    length: float = 1.18
    pitch: float = 0.0
    silence: float = 0.28
    volume: float = 1.0
    rms: float = 0.10  # RMS cible après synth, 0–1 (float)


# Narrateur ≠ papa ≠ enfant. Maîtresse ≠ maman. Copain ≠ héros.
# Tom était trop bas dans le mix (peak des autres voix) : volume + RMS + présence.
CAST = {
    "narrateur": Voice("fr_FR-tom-medium", length=1.22, pitch=0.2, silence=0.42, volume=1.55, rms=0.13),
    "maman": Voice("fr_FR-siwis-medium", length=1.20, pitch=0.0, silence=0.36, volume=1.05, rms=0.10),
    "papa": Voice("fr_FR-upmc-medium", speaker=1, length=1.18, pitch=0.0, silence=0.36, volume=1.12, rms=0.11),
    "maitresse": Voice("fr_FR-upmc-medium", speaker=0, length=1.16, pitch=-0.2, silence=0.34, volume=1.08, rms=0.10),
    "directrice": Voice("fr_FR-upmc-medium", speaker=0, length=1.18, pitch=-0.6, silence=0.36, volume=1.08, rms=0.10),
    "directeur": Voice("fr_FR-gilles-low", length=1.16, pitch=0.8, silence=0.36, volume=1.25, rms=0.11),
    "grand-mere": Voice("fr_FR-siwis-medium", length=1.32, pitch=-1.8, silence=0.40, volume=1.10, rms=0.10),
    "grand-pere": Voice("fr_FR-gilles-low", length=1.30, pitch=-0.6, silence=0.40, volume=1.25, rms=0.11),
    "nounou": Voice("fr_FR-upmc-medium", speaker=0, length=1.18, pitch=0.4, silence=0.36, volume=1.08, rms=0.10),
    "enfant-f": Voice("fr_FR-siwis-medium", length=1.28, pitch=2.4, silence=0.40, volume=1.08, rms=0.10),
    "enfant-m": Voice("fr_FR-tom-medium", length=1.28, pitch=2.6, silence=0.40, volume=1.20, rms=0.11),
    "copine": Voice("fr_FR-upmc-medium", speaker=0, length=1.24, pitch=2.2, silence=0.38, volume=1.08, rms=0.10),
    "copain": Voice("fr_FR-upmc-medium", speaker=1, length=1.24, pitch=2.8, silence=0.38, volume=1.12, rms=0.10),
}


def find_engine():
    piper = shutil.which("piper") or str(Path.home() / ".local/bin/piper")
    if piper and not Path(piper).exists():
        piper = None
    model = VOICES / "fr_FR-siwis-medium.onnx"
    if piper and model.exists():
        return "piper", piper, model
    espeak = shutil.which("espeak-ng") or shutil.which("espeak")
    if espeak:
        return "espeak", espeak, None
    return None, None, None


def pitch_shift(samples: np.ndarray, semitones: float) -> np.ndarray:
    if abs(semitones) < 0.05 or len(samples) < 32:
        return samples
    factor = 2.0 ** (semitones / 12.0)
    n = max(32, int(round(len(samples) / factor)))
    return resample(samples, n).astype(np.float32)


def match_loudness(samples: np.ndarray, target_rms: float) -> np.ndarray:
    """Aligne le RMS (voix Tom trop basse après un peak-norm du mix)."""
    if samples.size < 32:
        return samples
    rms = float(np.sqrt(np.mean(samples.astype(np.float64) ** 2)))
    if rms < 1.0:
        return samples
    target = target_rms * 32767.0
    out = samples * (target / rms)
    peak = float(np.max(np.abs(out)))
    cap = 0.95 * 32767.0
    if peak > cap:
        out *= cap / peak
    return out.astype(np.float32)


def presence_boost(samples: np.ndarray, sr: int, db: float = 3.5) -> np.ndarray:
    """Léger coup de présence (2–4 kHz) pour une voix d’homme trop sourde."""
    if samples.size < 64 or abs(db) < 0.2:
        return samples
    from scipy.signal import butter, sosfilt

    lo = min(2000 / (sr / 2), 0.45)
    hi = min(4200 / (sr / 2), 0.48)
    if hi <= lo:
        return samples
    sos = butter(2, [lo, hi], btype="band", output="sos")
    band = sosfilt(sos, samples)
    gain = 10 ** (db / 20.0) - 1.0
    return (samples + gain * band).astype(np.float32)


def read_pcm16(path: Path) -> tuple[np.ndarray, int]:
    with wave.open(str(path), "rb") as w:
        sr = w.getframerate()
        nch = w.getnchannels()
        sw = w.getsampwidth()
        n = w.getnframes()
        raw = w.readframes(n)
    if sw != 2:
        raise ValueError(f"{path}: sampwidth {sw}, attendu 16-bit")
    samples = np.frombuffer(raw, dtype="<i2").astype(np.float32)
    if nch > 1:
        samples = samples.reshape(-1, nch).mean(axis=1)
    return samples, sr


def to_stereo_i16(samples: np.ndarray) -> np.ndarray:
    mono = np.clip(samples, -32768, 32767).astype("<i2")
    stereo = np.empty((mono.shape[0], 2), dtype="<i2")
    stereo[:, 0] = mono
    stereo[:, 1] = mono
    return stereo


def write_wav44100(path: Path, samples: np.ndarray):
    path.parent.mkdir(parents=True, exist_ok=True)
    stereo = to_stereo_i16(samples)
    with wave.open(str(path), "wb") as w:
        w.setnchannels(CHANNELS)
        w.setsampwidth(2)
        w.setframerate(TARGET_SR)
        w.writeframes(stereo.tobytes())


def write_mp3(path: Path, samples: np.ndarray):
    import lameenc

    enc = lameenc.Encoder()
    enc.set_bit_rate(128)
    enc.set_in_sample_rate(TARGET_SR)
    enc.set_channels(CHANNELS)
    enc.set_quality(2)
    pcm = to_stereo_i16(samples).tobytes()
    mp3 = enc.encode(pcm) + enc.flush()
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(mp3)


def finalize(samples: np.ndarray, sr: int) -> tuple[np.ndarray, dict]:
    if sr != TARGET_SR:
        # 22050 → 44100 = *2 exact
        g = np.gcd(sr, TARGET_SR)
        samples = resample_poly(samples, TARGET_SR // g, sr // g)
    peak = float(np.max(np.abs(samples))) if len(samples) else 0.0
    if peak < MIN_PEAK:
        raise ValueError(f"signal trop faible peak={peak}")
    samples = samples * (PEAK_TARGET * 32767.0 / peak)
    rms = float(np.sqrt(np.mean(samples**2)))
    return samples, {
        "peak": int(np.max(np.abs(samples))),
        "rms": round(rms, 1),
        "dur": round(len(samples) / TARGET_SR, 2),
        "src_peak": int(peak),
    }


def stats_file(path: Path) -> dict:
    with wave.open(str(path), "rb") as w:
        sr = w.getframerate()
        nch = w.getnchannels()
        n = w.getnframes()
        raw = w.readframes(n)
    samples = np.frombuffer(raw, dtype="<i2").astype(np.float32)
    peak = int(np.max(np.abs(samples))) if len(samples) else 0
    rms = float(np.sqrt(np.mean(samples**2))) if len(samples) else 0
    return {
        "sr": sr,
        "ch": nch,
        "peak": peak,
        "rms": round(rms, 1),
        "dur": round(n / sr, 2) if sr else 0,
        "ok": sr == TARGET_SR and nch == CHANNELS and peak >= MIN_PEAK,
    }


def synth_piper(
    piper: str,
    model: Path,
    text: str,
    out: Path,
    length_scale: float,
    speaker: int | None = None,
    volume: float = 1.0,
):
    out.parent.mkdir(parents=True, exist_ok=True)
    cmd = [
        piper,
        "--model",
        str(model),
        "--output_file",
        str(out),
        "--length_scale",
        str(length_scale or 1.2),
        "--sentence_silence",
        "0.22",
        "--volume",
        str(volume or 1.0),
    ]
    if speaker is not None:
        cmd.extend(["--speaker", str(speaker)])
    subprocess.run(
        cmd,
        input=text.encode("utf-8"),
        check=True,
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
    )


def model_file(name: str) -> Path:
    p = VOICES / f"{name}.onnx"
    if p.exists():
        return p
    return VOICES / "fr_FR-siwis-medium.onnx"


FX_DIR = ROOT / "outils" / "fx"


def parse_sons(value) -> list[str]:
    if not value:
        return []
    return [s.strip() for s in str(value).split(",") if s.strip()]


def interleave_fx(beats: list[tuple[str, str]], sons: list[str]) -> list[tuple[str, str]]:
    """Après le premier récit : le bruit, puis la suite au calme. Jamais sous la voix."""
    if not sons:
        return beats
    fx = [("fx", s) for s in sons]
    if not beats:
        return fx
    out: list[tuple[str, str]] = []
    placed = False
    for i, beat in enumerate(beats):
        out.append(beat)
        if not placed and (beat[0] == "narrateur" or i == 0):
            out.extend(fx)
            placed = True
    if not placed:
        out = fx + beats
    return out


def load_fx_samples(son_id: str, sr: int) -> np.ndarray:
    path = FX_DIR / f"{son_id}.wav"
    if path.exists():
        samples, fsr = read_pcm16(path)
        if fsr != sr and len(samples):
            g = np.gcd(fsr, sr)
            samples = resample_poly(samples, sr // g, fsr // g)
        return samples.astype(np.float32)
    return np.zeros(int(sr * 1.15), dtype=np.float32)


def synth_beats(piper: str, beats: list[tuple[str, str]], td: Path) -> tuple[np.ndarray, int]:
    parts: list[np.ndarray] = []
    sr = 22050
    for i, (role, phrase) in enumerate(beats):
        if role == "fx":
            samples = load_fx_samples(phrase, sr)
            parts.append(samples)
            parts.append(np.zeros(int(sr * 0.35), dtype=np.float32))
            continue
        v = CAST.get(role, CAST["narrateur"])
        raw = td / f"{i}.wav"
        synth_piper(piper, model_file(v.model), phrase, raw, v.length, v.speaker, v.volume)
        samples, sr = read_pcm16(raw)
        samples = pitch_shift(samples, v.pitch)
        if role == "narrateur":
            samples = presence_boost(samples, sr, db=4.0)
        samples = match_loudness(samples, v.rms)
        parts.append(samples)
        parts.append(np.zeros(int(sr * (v.silence or GAP_S)), dtype=np.float32))
    if not parts:
        raise ValueError("aucune réplique")
    return np.concatenate(parts), sr


def synth_espeak(bin: str, text: str, out: Path, row: dict):
    out.parent.mkdir(parents=True, exist_ok=True)
    speed = max(80, min(160, int(row.get("rate_wpm") or 130)))
    pitch = int(row.get("espeak_pitch") or 50)
    gap = int(row.get("espeak_word_gap") or 10)
    amp = int(row.get("espeak_amp") or 100)
    subprocess.run(
        [
            bin, "-v", "fr", "-s", str(speed), "-p", str(pitch),
            "-g", str(gap), "-a", str(amp), "-w", str(out), text,
        ],
        check=True,
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
    )


def load_tree(xlsx: Path) -> tuple[dict, list[dict]]:
    wb = load_workbook(xlsx, read_only=True, data_only=True)
    meta: dict[str, str] = {}
    if "meta" in wb.sheetnames:
        for row in wb["meta"].iter_rows(values_only=True):
            if row and row[0] not in (None, "clé"):
                meta[str(row[0])] = "" if row[1] is None else str(row[1])
    ch = wb["chunks"]
    rows = list(ch.iter_rows(values_only=True))
    headers = [str(h) if h else "" for h in rows[0]]
    out = []
    roster = parse_roster(meta.get("characters", ""))
    for r in rows[1:]:
        d = {headers[i]: r[i] if i < len(r) else None for i in range(len(headers))}
        if not d.get("chunk_id") or not d.get("text"):
            continue
        if d.get("script"):
            d["beats"] = script_load(str(d["script"]))
        else:
            d["beats"] = split_script(str(d["text"]), roster)
        out.append(d)
    wb.close()
    return meta, out


def emit_playable(src_wav: Path, dest_wav: Path, dest_mp3: Path) -> dict:
    samples, sr = read_pcm16(src_wav)
    samples, info = finalize(samples, sr)
    write_wav44100(dest_wav, samples)
    write_mp3(dest_mp3, samples)
    check = stats_file(dest_wav)
    if not check["ok"] or check.get("ch", 1) != CHANNELS:
        raise RuntimeError(f"post-verif fail {dest_wav} {check}")
    if dest_mp3.stat().st_size < 400:
        raise RuntimeError(f"mp3 trop petit {dest_mp3}")
    info.update(check)
    return info


def bake_tree(xlsx: Path, engine, bin_path, model, force=False) -> tuple[int, int, int]:
    story_id = xlsx.stem
    dest = AUDIO / story_id
    _meta, chunks = load_tree(xlsx)
    ok = skip = fail = 0
    for d in chunks:
        cid = str(d["chunk_id"]).replace("/", "_")
        wav = dest / f"{cid}.wav"
        mp3 = dest / f"{cid}.mp3"
        if wav.exists() and mp3.exists() and not force:
            try:
                if stats_file(wav)["ok"]:
                    skip += 1
                    continue
            except Exception:
                pass
        beats = d.get("beats") or [("narrateur", str(d["text"]).strip())]
        beats = interleave_fx(beats, parse_sons(d.get("sons")))
        try:
            with tempfile.TemporaryDirectory() as td:
                tdir = Path(td)
                raw = tdir / "mix.wav"
                if engine == "piper":
                    samples, sr = synth_beats(bin_path, beats, tdir)
                    write_raw_pcm(raw, samples, sr)
                    emit_playable(raw, wav, mp3)
                else:
                    synth_espeak(bin_path, " ".join(p for _, p in beats), raw, d)
                    emit_playable(raw, wav, mp3)
            ok += 1
        except Exception as e:
            fail += 1
            print(f"FAIL {story_id}/{cid}: {e}", file=sys.stderr)
    return ok, skip, fail


def write_raw_pcm(path: Path, samples: np.ndarray, sr: int) -> None:
    mono = np.clip(samples, -32768, 32767).astype("<i2")
    with wave.open(str(path), "wb") as w:
        w.setnchannels(1)
        w.setsampwidth(2)
        w.setframerate(sr)
        w.writeframes(mono.tobytes())


def fix_existing(force=False) -> tuple[int, int]:
    """Réécrit les WAV 22050 déjà là en 44100+MP3, sans Piper."""
    ok = fail = 0
    for wav in sorted(AUDIO.rglob("*.wav")):
        if wav.with_suffix(".mp3").exists() and not force:
            try:
                if stats_file(wav)["ok"]:
                    continue
            except Exception:
                pass
        tmp = wav.with_suffix(".wav.tmp")
        try:
            info = emit_playable(wav, tmp, wav.with_suffix(".mp3"))
            tmp.replace(wav)
            ok += 1
            print(f"FIX {wav.parent.name}/{wav.name} sr44100 peak={info['peak']} dur={info['dur']}")
        except Exception as e:
            if tmp.exists():
                tmp.unlink()
            fail += 1
            print(f"FAIL fix {wav}: {e}", file=sys.stderr)
    return ok, fail


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--only", nargs="*", help="tree_id sans .xlsx")
    ap.add_argument("--limit", type=int, default=0)
    ap.add_argument("--force", action="store_true")
    ap.add_argument("--fix-existing", action="store_true", help="réécrire 22050→44100+mp3")
    ap.add_argument("--verify", action="store_true")
    args = ap.parse_args()

    if args.fix_existing:
        ok, fail = fix_existing(args.force)
        print(f"FIX DONE ok={ok} fail={fail}")
        if args.verify or True:
            bad = 0
            n = 0
            for wav in AUDIO.rglob("*.wav"):
                n += 1
                s = stats_file(wav)
                mp3 = wav.with_suffix(".mp3")
                if not s["ok"] or not mp3.exists():
                    bad += 1
                    print("BAD", wav, s, "mp3", mp3.exists())
            print(f"VERIFY n={n} bad={bad}")
        sys.exit(1 if fail else 0)

    engine, bin_path, model = find_engine()
    if not engine:
        print("Aucun TTS : installer piper ou espeak-ng", file=sys.stderr)
        sys.exit(1)
    print(f"engine={engine} bin={bin_path} model={model} out=wav44100-stereo+mp3-128k")
    files = sorted(ARBRES.glob("*.xlsx"))
    if args.only:
        want = set(args.only)
        files = [f for f in files if f.stem in want]
    if args.limit:
        files = files[: args.limit]
    total_ok = total_skip = total_fail = 0
    for i, f in enumerate(files, 1):
        ok, skip, fail = bake_tree(f, engine, bin_path, model, args.force)
        total_ok += ok
        total_skip += skip
        total_fail += fail
        print(f"[{i}/{len(files)}] {f.stem} +{ok} skip={skip} fail={fail}", flush=True)
    print(f"DONE ok={total_ok} skip={total_skip} fail={total_fail}")
    sys.exit(1 if total_fail else 0)


if __name__ == "__main__":
    main()
