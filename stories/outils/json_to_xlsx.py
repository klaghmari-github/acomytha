#!/usr/bin/env python3
"""JSON Sentier → un .xlsx par arbre. Décisions : stories/DECISIONS_EXCEL.md"""
from __future__ import annotations

import json
import re
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "arbres"
LECONS_XLSX = ROOT / "referentiel" / "lecons.xlsx"
LECONS_PATH = ROOT / "referentiel" / "lecons.json"  # archive, si xlsx absent

KINDS = (
    "passage_debut",
    "passage",
    "passage_question",
    "transition_question",
    "passage_fin",
)

CHUNK_COLS = [
    "chunk_id",
    "kind",
    "source_node_id",
    "lesson_id",
    "text",
    "text_ssml",
    "text_xai_tags",
    "expected_answer",
    "accepted_examples",
    "engine_ok_text",
    "engine_near_text",
    "retry_prompt",
    "option_1_label",
    "option_1_next_chunk",
    "option_2_label",
    "option_2_next_chunk",
    "option_3_label",
    "option_3_next_chunk",
    "default_next_chunk",
    "wait_ms",
    "retry_once",
    "night_policy",
    "locale",
    "voice_id",
    "rate_wpm",
    "rate_label",
    "speed_xai",
    "length_scale_piper",
    "pitch_label",
    "pitch_ssml",
    "pitch_xai_tag",
    "volume_label",
    "volume_db",
    "emphasis_words",
    "pause_before_ms",
    "pause_after_ms",
    "pause_sentence_ms",
    "style_energy",
    "style_contour",
    "noise_scale_piper",
    "ipa_replace",
    "kokoro_voice",
    "kokoro_speed",
    "melo_speed",
    "espeak_amp",
    "espeak_pitch",
    "espeak_word_gap",
    "notes",
]


def load_lecons():
    if LECONS_XLSX.exists():
        from openpyxl import load_workbook

        wb = load_workbook(LECONS_XLSX, read_only=True, data_only=True)
        rows = list(wb["lecons"].iter_rows(values_only=True))
        headers = [str(h) for h in rows[0]]
        out = {}
        for r in rows[1:]:
            d = {headers[i]: r[i] for i in range(len(headers))}
            lid = d.get("lesson_id")
            if not lid:
                continue
            for k in (
                "required_messages",
                "safe_actions",
                "misconceptions",
                "forbidden_in_audio",
                "answer_intents",
                "compatible_lessons",
            ):
                v = d.get(k) or ""
                d[k] = [x.strip() for x in str(v).split("|") if x.strip()]
            out[str(lid)] = d
        wb.close()
        return out
    if not LECONS_PATH.exists():
        return {}
    data = json.loads(LECONS_PATH.read_text(encoding="utf-8"))
    return {x["lesson_id"]: x for x in data.get("lessons", [])}


def clean_text(s: str) -> str:
    if not s:
        return ""
    s = s.replace("\u00a0", " ").replace("\n", " ")
    s = re.sub(r"\s+", " ", s).strip()
    s = s.replace("..", ".")
    if s and s[-1] not in ".!?":
        s += "."
    return s


def split_ending(text: str) -> tuple[str, str]:
    t = clean_text(text)
    fin = "L'histoire est finie."
    if fin.lower() in t.lower():
        body = re.split(r"L['']histoire est finie\.?", t, flags=re.I)[0].strip()
        if body and body[-1] not in ".!?":
            body += "."
        return body, fin
    return t, fin


def choice_prompt(labels: list[str], original: str) -> tuple[str, str]:
    labs = [clean_text(x).rstrip(".") for x in labels]
    if len(labs) == 1:
        neu = f"On peut prendre {labs[0]}."
    elif len(labs) == 2:
        neu = f"On peut prendre {labs[0]}, ou {labs[1]}."
    else:
        neu = f"On peut prendre {labs[0]}, {labs[1]}, ou {labs[2]}."
    note = ""
    orig = (original or "").lower()
    if any(x not in orig for x in labs[:1]) or "couleur" in orig:
        note = "prompt réécrit pour coller aux options"
    return neu, note


def listen_q(lesson: dict | None, examples: list[str], original: str) -> str:
    orig = clean_text(original)
    if orig and len(orig.split()) <= 18 and "?" in orig:
        return orig
    if examples:
        mot = examples[0]
        return f"On fait quoi ? {mot.capitalize()} ?"
    if lesson:
        sa = (lesson.get("safe_actions") or ["on se souvient"])[0]
        return f"On fait quoi ? {sa.capitalize()} ?"
    return "On se souvient ? On dit le petit mot ?"


def tts_profile(kind: str, age: str, framing: str) -> dict:
    n1 = age == "N1"
    n3 = age == "N3"
    critical = framing == "positive_only_critical"
    wpm = 118 if n1 else (148 if n3 else 132)
    length = 1.28 if n1 else (1.12 if n3 else 1.18)
    speed = 0.86 if n1 else (0.96 if n3 else 0.92)
    energy = "calm"
    pitch = "medium"
    contour = "level"
    pause_s = 450
    if kind == "passage_question":
        wpm -= 18
        length += 0.12
        speed -= 0.06
        energy = "warm"
        contour = "rise"
        pause_s = 250
    elif kind == "transition_question":
        wpm -= 15
        length += 0.10
        speed -= 0.05
        energy = "warm"
        contour = "rise"
        pause_s = 500
    elif kind == "passage_fin":
        wpm -= 12
        length += 0.08
        speed -= 0.04
        pitch = "low"
        energy = "calm"
        contour = "fall"
        pause_s = 600
    elif kind == "passage_debut":
        pause_s = 400
    if critical and kind in ("passage_question", "passage_fin", "passage"):
        wpm = min(wpm, 115)
        length = max(length, 1.22)
        pitch = "low"
    rate_label = "slow" if wpm < 125 else "medium"
    return {
        "locale": "fr-FR",
        "voice_id": "fr_FR-siwis-medium",
        "rate_wpm": wpm,
        "rate_label": rate_label,
        "speed_xai": round(speed, 2),
        "length_scale_piper": round(length, 2),
        "pitch_label": pitch,
        "pitch_ssml": "low" if pitch == "low" else "medium",
        "pitch_xai_tag": "lower-pitch" if pitch == "low" else "",
        "volume_label": "soft" if kind == "passage_fin" else "medium",
        "volume_db": -2 if kind == "passage_fin" else 0,
        "pause_before_ms": 200 if kind != "passage_debut" else 0,
        "pause_after_ms": pause_s,
        "pause_sentence_ms": 380 if n1 else 280,
        "style_energy": energy,
        "style_contour": contour,
        "noise_scale_piper": 0.333,
        "kokoro_voice": "ff_siwis",
        "kokoro_speed": round(speed, 2),
        "melo_speed": round(speed, 2),
        "espeak_amp": 80 if kind == "passage_fin" else 100,
        "espeak_pitch": 35 if pitch == "low" else 50,
        "espeak_word_gap": 12 if n1 else 8,
    }


def compile_ssml(text: str, prof: dict, emphasis: list[str]) -> str:
    body = text
    for w in emphasis:
        if w and w.lower() in body.lower():
            body = re.sub(
                re.escape(w),
                f'<emphasis level="moderate">{w}</emphasis>',
                body,
                count=1,
                flags=re.I,
            )
            break
    rate = prof["rate_label"]
    pitch = prof["pitch_ssml"]
    br = prof["pause_after_ms"]
    return (
        f'<speak><prosody rate="{rate}" pitch="{pitch}">{body}'
        f'</prosody><break time="{br}ms"/></speak>'
    )


def compile_xai(text: str, prof: dict, emphasis: list[str]) -> str:
    body = text
    for w in emphasis:
        if w and w.lower() in body.lower():
            body = re.sub(re.escape(w), f"<emphasis>{w}</emphasis>", body, count=1, flags=re.I)
            break
    inner = body
    if prof["rate_label"] == "slow":
        inner = f"<slow>{inner}</slow>"
    if prof["volume_label"] == "soft":
        inner = f"<soft>{inner}</soft>"
    if prof["pitch_xai_tag"]:
        inner = f"<{prof['pitch_xai_tag']}>{inner}</{prof['pitch_xai_tag']}>"
    pause = "[pause]" if prof["pause_after_ms"] >= 400 else ""
    if prof["pause_after_ms"] >= 800:
        pause = "[long-pause]"
    return f"{inner} {pause}".strip()


def emphasis_for(text: str, lesson: dict | None, framing: str) -> list[str]:
    keys = []
    if lesson:
        keys.extend(lesson.get("required_messages") or [])
        keys.extend(lesson.get("safe_actions") or [])
    keys.extend(["trottoir", "main", "stop", "assis", "attendre"])
    found = []
    low = text.lower()
    for k in keys:
        tok = k.split()[0] if k else ""
        if tok and tok.lower() in low and tok.lower() not in [x.lower() for x in found]:
            found.append(tok)
        if len(found) >= 1:
            break
    return found


class Builder:
    def __init__(self, tree: dict, lecons: dict):
        self.t = tree
        self.nodes = tree["nodes"]
        self.lecons = lecons
        self.lesson = lecons.get(tree.get("lesson_id") or "", {})
        self.rows: list[dict] = []
        self.journal: list[str] = []
        self.seen_nodes: set[str] = set()
        self.age = tree.get("age_band") or "N2"
        self.framing = tree.get("framing") or "standard"

    def add(self, **kw):
        kind = kw["kind"]
        text = clean_text(kw.get("text") or "")
        kw["text"] = text
        lesson = self.lecons.get(kw.get("lesson_id") or self.t.get("lesson_id") or "")
        emp = emphasis_for(text, lesson, self.framing)
        prof = tts_profile(kind, self.age, self.framing)
        kw.update(prof)
        kw["emphasis_words"] = ", ".join(emp)
        kw["text_ssml"] = compile_ssml(text, prof, emp)
        kw["text_xai_tags"] = compile_xai(text, prof, emp)
        kw.setdefault("ipa_replace", "")
        kw.setdefault("notes", "")
        kw.setdefault("wait_ms", 3000 if kind == "passage_question" else 0)
        kw.setdefault("retry_once", "oui" if kind == "passage_question" else "non")
        if kind == "passage_question":
            kw.setdefault("night_policy", "skip")
        elif kind == "transition_question":
            kw.setdefault("night_policy", "auto_default")
        else:
            kw.setdefault("night_policy", "play")
        kw.setdefault("engine_ok_text", "Oui, c'est la bonne réponse.")
        kw.setdefault("engine_near_text", "Tu étais presque.")
        self.rows.append(kw)

    def walk(self, nid: str, cid: str, t_depth: int, had_q: bool, is_debut: bool):
        if nid not in self.nodes:
            self.journal.append(f"nœud manquant {nid} depuis {cid}")
            return
        node = self.nodes[nid]
        ntype = node.get("type")
        if ntype == "silence_check":
            nxt = node.get("next") or node.get("default_next")
            if nxt:
                self.walk(nxt, cid, t_depth, had_q, is_debut)
            return

        if ntype in ("audio", "transition"):
            kind = "passage_debut" if is_debut else "passage"
            self.add(
                chunk_id=cid,
                kind=kind,
                source_node_id=nid,
                lesson_id=node.get("lesson_id") or self.t.get("lesson_id"),
                text=node.get("text") or "",
            )
            nxt = node.get("next")
            if nxt:
                self.dispatch(nxt, cid, t_depth, had_q, after_passage=True)
            return

        if ntype == "choice_story":
            self.emit_choice(nid, cid, t_depth, had_q)
            return

        if ntype in ("question_lesson", "question_comprehension"):
            self.emit_question(nid, cid, t_depth, True)
            return

        if ntype == "feedback":
            self.add(
                chunk_id=cid,
                kind="passage",
                source_node_id=nid,
                lesson_id=node.get("lesson_id") or self.t.get("lesson_id"),
                text=node.get("text") or "",
                notes="confirmation après question",
            )
            nxt = node.get("next")
            if nxt:
                self.dispatch(nxt, cid, t_depth, True, after_passage=True)
            return

        if ntype == "ending":
            self.emit_ending(nid, cid, t_depth, had_q)
            return

        self.journal.append(f"type ignoré {ntype} {nid}")

    def emit_choice(self, nid: str, cid: str, t_depth: int, had_q: bool):
        node = self.nodes[nid]
        opts = node.get("options") or []
        labels = [o.get("label") or o.get("intent") or f"option {i+1}" for i, o in enumerate(opts)]
        prompt, note = choice_prompt(labels, node.get("prompt") or "")
        if note:
            self.journal.append(f"{nid}: {note}")
        tnum = t_depth + 1
        stem = self.stem_of(cid)
        if stem == "CHK_T0000_P0000":
            tq_id = "CHK_T0001_P0000"
            tnum = 1
        else:
            tq_id = f"{stem}_T{tnum:04d}_P0000"

        next_ids = []
        for i, o in enumerate(opts[:3], start=1):
            next_ids.append(tq_id.replace("_P0000", f"_P{i:04d}"))
        while len(next_ids) < 3:
            next_ids.append("")
        default_i = 1
        dnext = node.get("default_next")
        for i, o in enumerate(opts[:3], start=1):
            if o.get("next") == dnext:
                default_i = i
        self.add(
            chunk_id=tq_id,
            kind="transition_question",
            source_node_id=nid,
            lesson_id=self.t.get("lesson_id"),
            text=prompt,
            option_1_label=labels[0] if labels else "",
            option_1_next_chunk=next_ids[0],
            option_2_label=labels[1] if len(labels) > 1 else "",
            option_2_next_chunk=next_ids[1],
            option_3_label=labels[2] if len(labels) > 2 else "",
            option_3_next_chunk=next_ids[2],
            default_next_chunk=next_ids[default_i - 1],
            wait_ms=3000,
            retry_once="oui",
            notes=note,
        )
        for i, o in enumerate(opts[:3], start=1):
            dest = o.get("next")
            child_cid = next_ids[i - 1]
            if dest:
                self.walk(dest, child_cid, tnum, had_q, is_debut=False)

    def stem_of(self, cid: str) -> str:
        cid = re.sub(r"(_Q0001|_C0001|_F0001|_END|_P)$", "", cid)
        return cid

    def emit_question(self, nid: str, cid: str, t_depth: int, had_q: bool):
        node = self.nodes[nid]
        stem = self.stem_of(cid)
        qid = f"{stem}_Q0001"
        conf_id = f"{stem}_C0001"
        examples = node.get("accepted_examples") or []
        qtxt = listen_q(self.lesson, examples, node.get("prompt") or "")
        nxt = node.get("next") or node.get("default_next")
        conf = node.get("positive_feedback") or "Oui. On se souvient du petit mot."
        if nxt and self.nodes.get(nxt, {}).get("type") == "feedback":
            fb = self.nodes[nxt]
            conf = fb.get("text") or conf
            nxt = fb.get("next")
        self.add(
            chunk_id=qid,
            kind="passage_question",
            source_node_id=nid,
            lesson_id=node.get("lesson_id") or self.t.get("lesson_id"),
            text=qtxt,
            expected_answer=(examples[0] if examples else ""),
            accepted_examples=" | ".join(examples),
            retry_prompt=clean_text(node.get("retry_prompt") or "Dis un petit mot."),
            default_next_chunk=conf_id,
            wait_ms=3000,
            retry_once="oui",
            notes="question d'écoute, ne change pas le cours",
        )
        self.add(
            chunk_id=conf_id,
            kind="passage",
            source_node_id=nid,
            lesson_id=node.get("lesson_id") or self.t.get("lesson_id"),
            text=conf,
            notes="confirmation ; le moteur peut dire engine_ok/near avant ce chunk",
        )
        if nxt:
            self.dispatch(nxt, stem, t_depth, True, after_passage=True)

    def emit_ending(self, nid: str, cid: str, t_depth: int, had_q: bool):
        node = self.nodes[nid]
        body, fin = split_ending(node.get("text") or "")
        if not had_q:
            qid = f"{cid}_Q0001"
            conf_id = f"{cid}_C0001"
            fin_id = f"{cid}_F0001"
            qtxt = listen_q(self.lesson, self.lesson.get("required_messages", []) if self.lesson else [], "")
            self.add(
                chunk_id=cid,
                kind="passage",
                source_node_id=nid,
                lesson_id=self.t.get("lesson_id"),
                text=body or fin,
            )
            self.add(
                chunk_id=qid,
                kind="passage_question",
                source_node_id=nid,
                lesson_id=self.t.get("lesson_id"),
                text=qtxt,
                expected_answer=(self.lesson.get("required_messages") or ["oui"])[0] if self.lesson else "oui",
                accepted_examples=" | ".join(self.lesson.get("required_messages") or []),
                default_next_chunk=conf_id,
                wait_ms=3000,
                retry_once="oui",
                notes="question d'écoute insérée avant la fin",
            )
            sa = ""
            if self.lesson:
                sa = (self.lesson.get("safe_actions") or [""])[0]
            self.add(
                chunk_id=conf_id,
                kind="passage",
                source_node_id=nid,
                lesson_id=self.t.get("lesson_id"),
                text=clean_text(f"Oui. {sa}." if sa else "Oui. On a bien fait."),
                notes="confirmation insérée",
            )
            self.add(
                chunk_id=fin_id,
                kind="passage_fin",
                source_node_id=nid,
                lesson_id=self.t.get("lesson_id"),
                text=fin,
            )
            self.journal.append(f"{nid}: question d'écoute insérée avant passage_fin")
            return
        # already had question: body as passage if non-empty, then fin
        if body and body != fin:
            self.add(
                chunk_id=cid,
                kind="passage",
                source_node_id=nid,
                lesson_id=self.t.get("lesson_id"),
                text=body,
            )
            self.add(
                chunk_id=f"{cid}_F0001",
                kind="passage_fin",
                source_node_id=nid,
                lesson_id=self.t.get("lesson_id"),
                text=fin,
            )
        else:
            self.add(
                chunk_id=cid,
                kind="passage_fin",
                source_node_id=nid,
                lesson_id=self.t.get("lesson_id"),
                text=fin,
            )

    def dispatch(self, nid: str, prev_cid: str, t_depth: int, had_q: bool, after_passage: bool):
        node = self.nodes.get(nid)
        if not node:
            return
        ntype = node.get("type")
        if ntype == "choice_story":
            self.emit_choice(nid, prev_cid, t_depth, had_q)
            return
        if ntype in ("question_lesson", "question_comprehension"):
            self.emit_question(nid, prev_cid, t_depth, had_q)
            return
        if ntype == "ending":
            self.emit_ending(nid, self.stem_of(prev_cid) + "_END", t_depth, had_q)
            return
        if ntype == "feedback":
            self.walk(nid, self.stem_of(prev_cid) + "_C0001", t_depth, True, is_debut=False)
            return
        if ntype in ("audio", "transition"):
            self.walk(nid, self.stem_of(prev_cid) + "_X", t_depth, had_q, is_debut=False)
            return
        if ntype == "silence_check":
            nxt = node.get("next") or node.get("default_next")
            if nxt:
                self.dispatch(nxt, prev_cid, t_depth, had_q, after_passage)
            return

    def build(self):
        root = self.t.get("root_id") or "root"
        self.walk(root, "CHK_T0000_P0000", 0, False, True)
        # unique chunk_ids
        seen = {}
        for r in self.rows:
            cid = r["chunk_id"]
            if cid in seen:
                seen[cid] += 1
                r["chunk_id"] = f"{cid}_x{seen[cid]}"
                self.journal.append(f"id dupliqué renommé {r['chunk_id']}")
            else:
                seen[cid] = 1


def write_xlsx(tree: dict, rows: list[dict], journal: list[str], dest: Path):
    wb = Workbook()
    meta = wb.active
    meta.title = "meta"
    t = tree
    seconds = t.get("duration_estimate_seconds") or {}
    meta_rows = [
        ("story_id", t.get("tree_id")),
        ("editorial_id", t.get("tree_id")),
        ("kind", t.get("kind")),
        ("title", t.get("title")),
        ("age_band", t.get("age_band")),
        ("age_range", t.get("age_range")),
        ("language", t.get("language") or "fr"),
        ("lesson_id", t.get("lesson_id")),
        ("secondary_lessons", ", ".join(t.get("secondary_lessons") or [])),
        ("domain", t.get("domain")),
        ("subdomain", t.get("subdomain")),
        ("framing", t.get("framing")),
        ("sensitivity", t.get("sensitivity")),
        ("family_model", t.get("family_model")),
        ("setting", t.get("setting")),
        ("characters", ", ".join(t.get("characters") or [])),
        ("duration_min_s", seconds.get("min")),
        ("duration_avg_s", seconds.get("avg")),
        ("duration_max_s", seconds.get("max")),
        ("engine_ok_text", "Oui, c'est la bonne réponse."),
        ("engine_near_text", "Tu étais presque."),
        ("engine_timeout_text", "On continue."),
        ("wait_default_ms", 3000),
        ("retry_once_default", "oui"),
        ("tts_engine_bake", "piper"),
        ("voice_id", "fr_FR-siwis-medium"),
        ("chunk_count", len(rows)),
        ("format_version", "xlsx-arbre-v1"),
    ]
    meta["A1"] = "clé"
    meta["B1"] = "valeur"
    meta["A1"].font = meta["B1"].font = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="2F5D50")
    meta["A1"].fill = meta["B1"].fill = fill
    for i, (k, v) in enumerate(meta_rows, start=2):
        meta[f"A{i}"] = k
        meta[f"B{i}"] = v if v is not None else ""
    meta.column_dimensions["A"].width = 28
    meta.column_dimensions["B"].width = 80

    ch = wb.create_sheet("chunks")
    header_fill = PatternFill("solid", fgColor="2F5D50")
    header_font = Font(bold=True, color="FFFFFF")
    wrap = Alignment(wrap_text=True, vertical="top")
    thin = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )
    for c, name in enumerate(CHUNK_COLS, start=1):
        cell = ch.cell(1, c, name)
        cell.fill = header_fill
        cell.font = header_font
    for r_i, row in enumerate(rows, start=2):
        for c, name in enumerate(CHUNK_COLS, start=1):
            val = row.get(name, "")
            if val is None:
                val = ""
            cell = ch.cell(r_i, c, val)
            cell.alignment = wrap
            cell.border = thin
            if name == "kind":
                colors = {
                    "passage_debut": "C5E0D5",
                    "passage": "EEF6F2",
                    "passage_question": "F7E3B0",
                    "transition_question": "D6E4F0",
                    "passage_fin": "E8D5F2",
                }
                cell.fill = PatternFill("solid", fgColor=colors.get(str(val), "FFFFFF"))
    dv = DataValidation(type="list", formula1='"' + ",".join(KINDS) + '"', allow_blank=False)
    ch.add_data_validation(dv)
    dv.add(f"B2:B{len(rows)+1}")
    ch.freeze_panes = "A2"
    ch.auto_filter.ref = f"A1:{get_column_letter(len(CHUNK_COLS))}{len(rows)+1}"
    widths = {
        "chunk_id": 42,
        "kind": 22,
        "text": 70,
        "text_ssml": 50,
        "text_xai_tags": 50,
        "notes": 36,
    }
    for c, name in enumerate(CHUNK_COLS, start=1):
        ch.column_dimensions[get_column_letter(c)].width = widths.get(name, 18)
    ch.row_dimensions[1].height = 22

    jo = wb.create_sheet("journal")
    jo["A1"] = "conversion / optimisations"
    jo["A1"].font = Font(bold=True)
    if not journal:
        journal = ["conversion structurelle, prompts de choix alignés, TTS profilé"]
    for i, line in enumerate(journal, start=2):
        jo[f"A{i}"] = line
    jo.column_dimensions["A"].width = 100

    leg = wb.create_sheet("legend")
    legend = [
        ("kind", "rôle"),
        ("passage_debut", "premier récit, CHK_T0000_P0000"),
        ("passage", "scène ou confirmation après question d'écoute"),
        ("passage_question", "question qui NE branche PAS ; wait_ms puis suite"),
        ("transition_question", "choix narratif 2 ou 3 options, toutes sûres"),
        ("passage_fin", "fermeture, contient « L'histoire est finie. »"),
        ("wait_ms", "silence moteur (pas dans le MP3). Défaut 3000"),
        ("engine_ok_text", "phrase moteur future si similarité haute"),
        ("engine_near_text", "phrase moteur future si similarité moyenne"),
        ("text", "français enfant, SANS balises"),
        ("text_ssml", "export SSML 1.1"),
        ("text_xai_tags", "export balises xAI (non utilisé au bake 0€)"),
        ("length_scale_piper", "vitesse Piper (bake actuel)"),
        ("speed_xai", "si un jour API TTS"),
        ("night_policy", "play | skip | auto_default"),
    ]
    for i, pair in enumerate(legend, start=1):
        leg.cell(i, 1, pair[0])
        leg.cell(i, 2, pair[1])
    leg.column_dimensions["A"].width = 24
    leg.column_dimensions["B"].width = 70

    dest.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dest)


def convert_one(path: Path, lecons: dict, out_dir: Path) -> tuple[str, int, list[str]]:
    tree = json.loads(path.read_text(encoding="utf-8"))
    b = Builder(tree, lecons)
    b.build()
    dest = out_dir / f"{tree['tree_id']}.xlsx"
    write_xlsx(tree, b.rows, b.journal, dest)
    return tree["tree_id"], len(b.rows), b.journal


def main():
    lecons = load_lecons()
    files = sorted(ROOT.glob("atomiques/**/*.json")) + sorted(ROOT.glob("ramifiees/**/*.json"))
    if not files:
        print("aucun JSON", file=sys.stderr)
        sys.exit(1)
    OUT.mkdir(parents=True, exist_ok=True)
    n = 0
    chunks = 0
    errors = []
    for p in files:
        try:
            tid, nc, _ = convert_one(p, lecons, OUT)
            n += 1
            chunks += nc
            if n % 100 == 0:
                print(f"... {n} arbres, {chunks} chunks", flush=True)
        except Exception as e:
            errors.append(f"{p}: {e}")
    print(f"OK {n} xlsx → {OUT}  chunks={chunks}  erreurs={len(errors)}")
    for e in errors[:30]:
        print("ERR", e)
    if errors:
        sys.exit(2)


if __name__ == "__main__":
    main()
