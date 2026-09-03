#!/usr/bin/env python3
"""lecons.json + arbres/*.xlsx → referentiel Excel (catalogue + liaisons)."""
from __future__ import annotations

import json
import re
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
REF = ROOT / "referentiel"
ARBRES = ROOT / "arbres"
JSON_LECONS = REF / "lecons.json"

HDR = PatternFill("solid", fgColor="2F5D50")
HFONT = Font(bold=True, color="FFFFFF")
WRAP = Alignment(wrap_text=True, vertical="top")


def join_list(v) -> str:
    if v is None:
        return ""
    if isinstance(v, list):
        return " | ".join(str(x) for x in v)
    return str(v)


def header(ws, cols):
    for i, c in enumerate(cols, 1):
        cell = ws.cell(1, i, c)
        cell.fill = HDR
        cell.font = HFONT
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"


def autosize(ws, cols, widths=None):
    widths = widths or {}
    for i, c in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(c, min(28, max(12, len(c) + 2)))


def write_lecons_xlsx(lessons: list[dict], dest: Path):
    wb = Workbook()
    # --- domaines ---
    doms = {}
    subs = {}
    for L in lessons:
        doms[L["domain_id"]] = L["domain"]
        subs[L["subdomain_id"]] = (L["domain_id"], L["subdomain"])

    dws = wb.active
    dws.title = "domaines"
    header(dws, ["domain_id", "domain", "n_lecons"])
    counts = {}
    for L in lessons:
        counts[L["domain_id"]] = counts.get(L["domain_id"], 0) + 1
    for i, (did, name) in enumerate(sorted(doms.items()), 2):
        dws.cell(i, 1, did)
        dws.cell(i, 2, name)
        dws.cell(i, 3, counts[did])
    autosize(dws, ["domain_id", "domain", "n_lecons"], {"domain": 28})

    sws = wb.create_sheet("sous_domaines")
    header(sws, ["subdomain_id", "subdomain", "domain_id", "n_lecons"])
    scount = {}
    for L in lessons:
        scount[L["subdomain_id"]] = scount.get(L["subdomain_id"], 0) + 1
    for i, (sid, (did, name)) in enumerate(sorted(subs.items()), 2):
        sws.cell(i, 1, sid)
        sws.cell(i, 2, name)
        sws.cell(i, 3, did)
        sws.cell(i, 4, scount[sid])
    autosize(sws, ["subdomain_id", "subdomain", "domain_id", "n_lecons"], {"subdomain": 32})

    cols = [
        "lesson_id",
        "title",
        "title_child_audio",
        "domain_id",
        "domain",
        "subdomain_id",
        "subdomain",
        "objective",
        "framing",
        "sensitivity",
        "wave",
        "family_model",
        "do_not_diagnose",
        "required_messages",
        "safe_actions",
        "misconceptions",
        "forbidden_in_audio",
        "answer_intents",
        "compatible_lessons",
    ]
    lws = wb.create_sheet("lecons")
    header(lws, cols)
    for r, L in enumerate(lessons, 2):
        vals = [
            L.get("lesson_id"),
            L.get("title"),
            L.get("title_child_audio"),
            L.get("domain_id"),
            L.get("domain"),
            L.get("subdomain_id"),
            L.get("subdomain"),
            L.get("objective"),
            L.get("framing"),
            L.get("sensitivity"),
            L.get("wave"),
            L.get("family_model"),
            "oui" if L.get("do_not_diagnose") else "non",
            join_list(L.get("required_messages")),
            join_list(L.get("safe_actions")),
            join_list(L.get("misconceptions")),
            join_list(L.get("forbidden_in_audio")),
            join_list(L.get("answer_intents")),
            join_list(L.get("compatible_lessons")),
        ]
        for c, v in enumerate(vals, 1):
            cell = lws.cell(r, c, v if v is not None else "")
            cell.alignment = WRAP
    lws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(lessons)+1}"
    autosize(
        lws,
        cols,
        {
            "lesson_id": 16,
            "title": 36,
            "title_child_audio": 28,
            "objective": 50,
            "required_messages": 40,
            "safe_actions": 40,
            "compatible_lessons": 40,
            "domain": 22,
            "subdomain": 22,
        },
    )

    # exploded attributes for filters
    ews = wb.create_sheet("attributs")
    ecols = ["lesson_id", "attr", "valeur"]
    header(ews, ecols)
    ri = 2
    mapping = [
        ("required_messages", "required_message"),
        ("safe_actions", "safe_action"),
        ("misconceptions", "misconception"),
        ("forbidden_in_audio", "forbidden"),
        ("answer_intents", "answer_intent"),
        ("compatible_lessons", "compatible_lesson"),
    ]
    for L in lessons:
        for src, attr in mapping:
            for val in L.get(src) or []:
                ews.cell(ri, 1, L["lesson_id"])
                ews.cell(ri, 2, attr)
                ews.cell(ri, 3, val)
                ri += 1
    ews.auto_filter.ref = f"A1:C{ri-1}"
    autosize(ews, ecols, {"valeur": 50})

    info = wb.create_sheet("meta")
    info["A1"] = "clé"
    info["B1"] = "valeur"
    info["A1"].fill = info["B1"].fill = HDR
    info["A1"].font = info["B1"].font = HFONT
    for i, (k, v) in enumerate(
        [
            ("document", "referentiel leçons AcoMytha"),
            ("n_lecons", len(lessons)),
            ("n_domaines", len(doms)),
            ("n_sous_domaines", len(subs)),
            ("source", "Source Unique v3.0 + Corrections v3.1"),
            ("format_version", "lecons-xlsx-v1"),
        ],
        2,
    ):
        info.cell(i, 1, k)
        info.cell(i, 2, v)

    dest.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dest)


def load_story_xlsx(path: Path) -> dict | None:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        meta = {}
        for row in wb["meta"].iter_rows(values_only=True):
            if row and row[0] and row[0] != "clé":
                meta[str(row[0])] = row[1]
        ch = wb["chunks"]
        rows = list(ch.iter_rows(values_only=True))
        headers = [str(h) if h else "" for h in rows[0]]
        chunks = []
        for r in rows[1:]:
            d = {headers[i]: (r[i] if i < len(r) else None) for i in range(len(headers))}
            if d.get("chunk_id"):
                chunks.append(d)
        return {"meta": meta, "chunks": chunks}
    finally:
        wb.close()


def specific_role(chunk: dict) -> str | None:
    """Un chunk est 'spécifique leçon' s'il porte la pédagogie, pas le simple récit de branche."""
    kind = chunk.get("kind") or ""
    cid = str(chunk.get("chunk_id") or "")
    if kind == "passage_question":
        return "question"
    if kind == "passage_fin":
        return "fin"
    if cid.endswith("_C0001") or "confirmation" in str(chunk.get("notes") or "").lower():
        return "confirmation"
    if kind == "passage_debut":
        return "accroche"
    return None


def text_hits_lesson(text: str, lesson: dict) -> bool:
    if not text:
        return False
    low = text.lower()
    needles = []
    for key in ("required_messages", "safe_actions"):
        for x in lesson.get(key) or []:
            tok = str(x).strip().lower()
            if len(tok) >= 4:
                needles.append(tok)
    return any(n in low for n in needles)


def build_links(lessons: list[dict], arbres: Path):
    by_id = {L["lesson_id"]: L for L in lessons}
    histoires = []  # lesson, story, role, ...
    chunks_out = []
    files = sorted(arbres.glob("*.xlsx"))
    n = 0
    for path in files:
        data = load_story_xlsx(path)
        n += 1
        if n % 200 == 0:
            print(f"  scan {n}/{len(files)}", flush=True)
        if not data:
            continue
        m = data["meta"]
        story_id = str(m.get("story_id") or path.stem)
        principal = str(m.get("lesson_id") or "").strip()
        secondary = [x.strip() for x in str(m.get("secondary_lessons") or "").split(",") if x.strip()]
        title = m.get("title") or ""
        kind = m.get("kind") or ""
        age = m.get("age_band") or ""
        linked = []
        if principal:
            linked.append((principal, "principal"))
        for s in secondary:
            if s != principal:
                linked.append((s, "secondaire"))
        for lid, role in linked:
            histoires.append(
                {
                    "lesson_id": lid,
                    "story_id": story_id,
                    "role": role,
                    "story_title": title,
                    "kind": kind,
                    "age_band": age,
                    "n_chunks_story": len(data["chunks"]),
                }
            )
        linked_ids = {x[0] for x in linked}
        for ch in data["chunks"]:
            role_c = specific_role(ch)
            if not role_c or not principal:
                continue
            text = str(ch.get("text") or "")
            assigned = str(ch.get("lesson_id") or "").strip() or principal
            if assigned not in linked_ids:
                assigned = principal
            for s, r in linked:
                if r == "secondaire" and s in by_id and text_hits_lesson(text, by_id[s]):
                    if principal not in by_id or not text_hits_lesson(text, by_id[principal]):
                        assigned = s
            chunks_out.append(
                {
                    "lesson_id": assigned,
                    "story_id": story_id,
                    "chunk_id": ch.get("chunk_id"),
                    "kind": ch.get("kind"),
                    "specifique": role_c,
                    "role_histoire": "principal" if assigned == principal else "secondaire",
                    "extrait": text[:180],
                }
            )
    return histoires, chunks_out


def write_links_xlsx(lessons, histoires, chunks_out, dest: Path):
    wb = Workbook()
    # couverture
    cov = {}
    for h in histoires:
        d = cov.setdefault(
            h["lesson_id"],
            {"principal": 0, "secondaire": 0, "stories": set(), "chunks": 0},
        )
        d[h["role"]] = d.get(h["role"], 0) + 1
        d["stories"].add(h["story_id"])
    for c in chunks_out:
        cov.setdefault(
            c["lesson_id"],
            {"principal": 0, "secondaire": 0, "stories": set(), "chunks": 0},
        )
        cov[c["lesson_id"]]["chunks"] += 1

    cws = wb.active
    cws.title = "couverture"
    ccols = [
        "lesson_id",
        "title",
        "domain_id",
        "subdomain_id",
        "n_histoires_principal",
        "n_histoires_secondaire",
        "n_histoires_total",
        "n_chunks_specifiques",
        "trou",
    ]
    header(cws, ccols)
    for i, L in enumerate(lessons, 2):
        lid = L["lesson_id"]
        d = cov.get(lid, {"principal": 0, "secondaire": 0, "stories": set(), "chunks": 0})
        ntot = len(d["stories"])
        trou = "oui" if d["principal"] == 0 else "non"
        vals = [
            lid,
            L.get("title"),
            L.get("domain_id"),
            L.get("subdomain_id"),
            d.get("principal", 0),
            d.get("secondaire", 0),
            ntot,
            d.get("chunks", 0),
            trou,
        ]
        for c, v in enumerate(vals, 1):
            cws.cell(i, c, v)
    cws.auto_filter.ref = f"A1:{get_column_letter(len(ccols))}{len(lessons)+1}"
    autosize(cws, ccols, {"title": 36, "lesson_id": 16})

    hws = wb.create_sheet("histoires")
    hcols = ["lesson_id", "story_id", "role", "story_title", "kind", "age_band", "n_chunks_story"]
    header(hws, hcols)
    histoires_sorted = sorted(histoires, key=lambda x: (x["lesson_id"], x["role"], x["story_id"]))
    for i, h in enumerate(histoires_sorted, 2):
        for c, k in enumerate(hcols, 1):
            cell = hws.cell(i, c, h.get(k, ""))
            cell.alignment = WRAP
    hws.auto_filter.ref = f"A1:{get_column_letter(len(hcols))}{len(histoires_sorted)+1}"
    autosize(hws, hcols, {"story_title": 50, "story_id": 28, "lesson_id": 16})

    kws = wb.create_sheet("chunks")
    kcols = ["lesson_id", "story_id", "chunk_id", "kind", "specifique", "role_histoire", "extrait"]
    header(kws, kcols)
    chunks_sorted = sorted(chunks_out, key=lambda x: (x["lesson_id"], x["story_id"], str(x["chunk_id"])))
    for i, ch in enumerate(chunks_sorted, 2):
        for c, k in enumerate(kcols, 1):
            cell = kws.cell(i, c, ch.get(k, ""))
            cell.alignment = WRAP
    if chunks_sorted:
        kws.auto_filter.ref = f"A1:{get_column_letter(len(kcols))}{len(chunks_sorted)+1}"
    autosize(kws, kcols, {"chunk_id": 48, "extrait": 60, "story_id": 28, "lesson_id": 16})

    info = wb.create_sheet("meta")
    info["A1"] = "clé"
    info["B1"] = "valeur"
    info["A1"].fill = info["B1"].fill = HDR
    info["A1"].font = info["B1"].font = HFONT
    for i, (k, v) in enumerate(
        [
            ("document", "liaison leçon → histoires → chunks pédagogiques"),
            ("n_liens_histoire", len(histoires)),
            ("n_chunks_specifiques", len(chunks_out)),
            ("specifique", "question | confirmation | fin | accroche"),
            ("format_version", "lecon-histoires-xlsx-v1"),
        ],
        2,
    ):
        info.cell(i, 1, k)
        info.cell(i, 2, v)

    dest.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dest)


def main():
    data = json.loads(JSON_LECONS.read_text(encoding="utf-8"))
    lessons = data["lessons"]
    p1 = REF / "lecons.xlsx"
    print("écriture", p1)
    write_lecons_xlsx(lessons, p1)
    print("scan", ARBRES)
    histoires, chunks_out = build_links(lessons, ARBRES)
    p2 = REF / "lecon_histoires.xlsx"
    print("écriture", p2, "liens", len(histoires), "chunks", len(chunks_out))
    write_links_xlsx(lessons, histoires, chunks_out, p2)
    print("OK")


if __name__ == "__main__":
    main()
