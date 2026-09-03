"""F-AUD-006 — scripts multi-voix et roster de rôles.

Le narrateur décrit. Les personnages parlent sans « maman dit ».
"""

from __future__ import annotations

import argparse
import json
import re
import unicodedata
from dataclasses import dataclass, field
from functools import lru_cache
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
ARBRES = ROOT / "arbres"
LEXIQUE = ROOT / "outils" / "fx" / "lexique.json"

FEMALE = {
    "adéle", "adèle", "agathe", "aline", "amandine", "anaïs", "aurore", "ava",
    "bérenice", "bérénice", "cécile", "céline", "clara", "coralie", "delphine",
    "diane", "dounia", "éléonore", "estelle", "éva", "faustine", "fatou",
    "flora", "flore", "gisèle", "hélène", "hortense", "inès", "iris", "jeanne",
    "léa", "lila", "lina", "maëlys", "maya", "nina", "ninon", "nora", "océane",
    "prune", "sara", "violette", "yasmine", "zoé",
}
MALE = {
    "achille", "adam", "amir", "armand", "baptiste", "bertrand", "brice",
    "césar", "clément", "côme", "damien", "denis", "didier", "dorian", "étienne",
    "ewen", "fabrice", "félix", "ferdinand", "florian", "gabin", "gaspard",
    "gaston", "hadrien", "hervé", "hippolyte", "honoré", "hugo", "idris",
    "idriss", "jules", "kenzo", "kilian", "loïc", "maël", "marceau", "nino",
    "noé", "octave", "sami", "théo", "tom", "ugo",
}
FAMILY = {
    "maman": "maman",
    "papa": "papa",
    "mamie": "grand-mere",
    "mémé": "grand-mere",
    "grand-mère": "grand-mere",
    "grand mere": "grand-mere",
    "papi": "grand-pere",
    "pépé": "grand-pere",
    "grand-père": "grand-pere",
    "grand pere": "grand-pere",
    "nounou": "nounou",
    "maîtresse": "maitresse",
    "maitresse": "maitresse",
    "la maîtresse": "maitresse",
    "la maitresse": "maitresse",
    "directrice": "directrice",
    "la directrice": "directrice",
    "directeur": "directeur",
    "le directeur": "directeur",
}
SKIP_WHO = {"il", "elle", "on", "i", "l"}

ATTRIB_RE = re.compile(
    r"(?P<who>la\s+ma[îi]tresse|le\s+directeur|la\s+directrice|"
    r"maman|papa|mamie|mémé|papi|pépé|grand-m[eè]re|grand-p[eè]re|"
    r"nounou|ma[îi]tresse|directrice|directeur|"
    r"[A-ZÉÈÊÀÂÎÔÛŸ][\w'\-]*)\s+dit\s*:\s*",
    re.IGNORECASE,
)


def _fold(s: str) -> str:
    s = unicodedata.normalize("NFD", s.lower().strip())
    return "".join(c for c in s if unicodedata.category(c) != "Mn")


def gender_of(name: str) -> str:
    n = _fold(name)
    if n in {_fold(x) for x in FEMALE}:
        return "f"
    if n in {_fold(x) for x in MALE}:
        return "m"
    if n.endswith(("a", "e", "ie", "ette", "ine", "elle")) and n not in {
        "noe",
        "andre",
        "pierre",
        "maxime",
        "felix",
        "come",
    }:
        return "f"
    return "m"


@dataclass
class Roster:
    hero: str = ""
    hero_role: str = "enfant-f"
    names: dict[str, str] = field(default_factory=dict)

    def resolve(self, who: str) -> str:
        key = who.lower().strip()
        key = re.sub(r"\s+", " ", key)
        if key in FAMILY:
            return FAMILY[key]
        if _fold(key) in {_fold(k) for k in FAMILY}:
            for k, v in FAMILY.items():
                if _fold(k) == _fold(key):
                    return v
        if _fold(key) in SKIP_WHO:
            return "narrateur"
        if key in self.names:
            return self.names[key]
        for n, role in self.names.items():
            if _fold(n) == _fold(key):
                return role
        return "enfant-f" if gender_of(who) == "f" else "enfant-m"


def parse_roster(characters: str) -> Roster:
    r = Roster()
    if not characters:
        return r
    parts = [p.strip() for p in characters.replace(" et ", ", ").split(",") if p.strip()]
    children: list[str] = []
    for p in parts:
        low = p.lower()
        if low in FAMILY or _fold(low) in {_fold(k) for k in FAMILY}:
            continue
        children.append(p)
    if not children:
        return r
    r.hero = children[0]
    r.hero_role = "enfant-f" if gender_of(r.hero) == "f" else "enfant-m"
    r.names[r.hero.lower()] = r.hero_role
    for other in children[1:]:
        g = gender_of(other)
        r.names[other.lower()] = "copine" if g == "f" else "copain"
    return r


def _cap(s: str) -> str:
    s = s.strip().strip("«»\"'“”")
    if not s:
        return s
    return s[0].upper() + s[1:]


def _clean_speech(s: str) -> str:
    s = s.strip()
    s = s.strip("«»\"'“”")
    s = re.sub(r"\s+", " ", s).strip()
    return _cap(s)


def split_script(text: str, roster: Roster) -> list[tuple[str, str]]:
    """Découpe un texte atelier en répliques (role, phrase)."""
    if not text or not str(text).strip():
        return []
    raw = str(text).replace("\n", " ").strip()
    raw = re.sub(r"\s+", " ", raw)
    matches = list(ATTRIB_RE.finditer(raw))
    if not matches:
        return [("narrateur", _cap(raw))]

    beats: list[tuple[str, str]] = []
    pos = 0
    for i, m in enumerate(matches):
        before = raw[pos : m.start()].strip()
        if before:
            beats.append(("narrateur", _cap(before)))
        who = m.group("who")
        end = matches[i + 1].start() if i + 1 < len(matches) else len(raw)
        spoken = raw[m.end() : end].strip()
        speech, narr = _split_speech_narr(spoken, roster)
        if speech:
            beats.append((roster.resolve(who), _clean_speech(speech)))
        if narr:
            beats.append(("narrateur", _cap(narr)))
        pos = end
    tail = raw[pos:].strip()
    if tail:
        beats.append(("narrateur", _cap(tail)))
    return [(a, b) for a, b in beats if b]


_NARR_START = re.compile(
    r"^(?:elle|il|ils|elles)\s+(?:prend|prennent|met|mets|regarde|sourit|marche|arrive|s'|se\s)|"
    r"^(?:une|un|le|la|les|dans|aujourd'hui|puis|alors)\b",
    re.IGNORECASE,
)


def _split_speech_narr(spoken: str, roster: Roster) -> tuple[str, str]:
    spoken = spoken.strip()
    if not spoken:
        return "", ""
    bits = re.split(r"(?<=[.!?])\s+", spoken)
    names = {_fold(roster.hero)} | {_fold(n) for n in roster.names}
    speech = [bits[0]]
    narr: list[str] = []
    mode = "speech"
    for s in bits[1:]:
        first = _fold(s.split(" ", 1)[0].strip(",. ")) if s else ""
        if first in names or _NARR_START.match(s or ""):
            mode = "narr"
        if mode == "narr":
            narr.append(s)
        else:
            speech.append(s)
    return " ".join(speech).strip(), " ".join(narr).strip()


def script_dump(beats: list[tuple[str, str]]) -> str:
    return "\n".join(f"{role}|{phrase}" for role, phrase in beats)


def script_load(blob: str) -> list[tuple[str, str]]:
    out = []
    for line in (blob or "").splitlines():
        if "|" not in line:
            continue
        role, phrase = line.split("|", 1)
        role, phrase = role.strip(), phrase.strip()
        if role and phrase:
            out.append((role, phrase))
    return out


def spoken_text(beats: list[tuple[str, str]]) -> str:
    return " ".join(p for _, p in beats)


@lru_cache(maxsize=1)
def load_lexique() -> list[dict]:
    if not LEXIQUE.exists():
        return []
    data = json.loads(LEXIQUE.read_text(encoding="utf-8"))
    return list(data.get("cues") or [])


def detect_sons(text: str) -> str:
    """Ids de sons pour un passage. Chaîne vide = silence (cas normal)."""
    if not text:
        return ""
    blob = _fold(str(text))
    found: list[str] = []
    for cue in load_lexique():
        cid = cue.get("id") or ""
        for pat in cue.get("match") or []:
            if _fold(pat) and _fold(pat) in blob:
                found.append(cid)
                break
    return ",".join(dict.fromkeys(found))


def fill_sons_xlsx(path: Path) -> tuple[int, int]:
    """Ajoute / remplit la colonne sons. Vide = passage silencieux."""
    wb = load_workbook(path)
    ws = wb["chunks"]
    headers = [c.value for c in ws[1]]
    if "sons" not in headers:
        ws.cell(1, len(headers) + 1, "sons")
        headers.append("sons")
    ti = headers.index("text") + 1 if "text" in headers else None
    si = headers.index("sons") + 1
    n_fx = n_quiet = 0
    for r in range(2, ws.max_row + 1):
        text = ws.cell(r, ti).value if ti else ""
        sons = detect_sons(str(text) if text else "")
        ws.cell(r, si, sons)
        if sons:
            n_fx += 1
        else:
            n_quiet += 1
    if "legend" in wb.sheetnames:
        lg = wb["legend"]
        already = any((c.value or "") == "sons" for c in lg["A"])
        if not already:
            lg.append(["sons", "ids de bruits (vide = silence). Le bruit se joue, puis l'histoire reprend au calme."])
    if "journal" in wb.sheetnames:
        wb["journal"].append(["F-AUD-007 colonne sons ; vide = silence ; jamais parler dans le bruit"])
    wb.save(path)
    wb.close()
    return n_fx, n_quiet


def rewrite_xlsx(path: Path) -> int:
    wb = load_workbook(path)
    meta = {}
    if "meta" in wb.sheetnames:
        for row in wb["meta"].iter_rows(values_only=True):
            if row and row[0] not in (None, "clé"):
                meta[str(row[0])] = "" if row[1] is None else str(row[1])
    roster = parse_roster(meta.get("characters", ""))
    ws = wb["chunks"]
    headers = [c.value for c in ws[1]]
    if "script" not in headers:
        col = len(headers) + 1
        ws.cell(1, col, "script")
        headers.append("script")
    ti = headers.index("text") + 1
    si = headers.index("script") + 1
    n = 0
    for r in range(2, ws.max_row + 1):
        text = ws.cell(r, ti).value
        if not text:
            continue
        beats = split_script(str(text), roster)
        if not beats:
            continue
        new_text = spoken_text(beats)
        ws.cell(r, ti, new_text)
        ws.cell(r, si, script_dump(beats))
        n += 1
    if "meta" in wb.sheetnames:
        ms = wb["meta"]
        keys = {str(c.value): c.row for c in ms["A"] if c.value}
        if "voice_cast" in keys:
            ms.cell(keys["voice_cast"], 2, "voix-roles-v1")
        else:
            ms.append(["voice_cast", "voix-roles-v1"])
    if "journal" in wb.sheetnames:
        js = wb["journal"]
        js.cell(js.max_row + 1, 1, "F-AUD-006 scripts multi-voix, sans « X dit »")
    wb.save(path)
    wb.close()
    return n


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--only", nargs="*")
    ap.add_argument("--limit", type=int, default=0)
    ap.add_argument("--sons-only", action="store_true", help="colonne sons seulement")
    args = ap.parse_args()
    files = sorted(ARBRES.glob("*.xlsx"))
    if args.only:
        want = set(args.only)
        files = [f for f in files if f.stem in want]
    if args.limit:
        files = files[: args.limit]
    total = fx = quiet = 0
    for i, f in enumerate(files, 1):
        if args.sons_only:
            a, b = fill_sons_xlsx(f)
            fx += a
            quiet += b
        else:
            total += rewrite_xlsx(f)
        if i % 50 == 0 or i == len(files):
            print(f"[{i}/{len(files)}] {f.stem}", flush=True)
    if args.sons_only:
        print(f"DONE files={len(files)} avec_sons={fx} silence={quiet}")
    else:
        print(f"DONE files={len(files)} chunks={total}")


if __name__ == "__main__":
    main()
