#!/usr/bin/env python3
"""avis1.txt — vraie histoire par chemin. Graphe et IDs inchangés."""

from __future__ import annotations

from openpyxl import load_workbook

XLSX = "stories/arbres/TREE-AUT-001.xlsx"


def S(pairs: list[tuple[str, str]]) -> tuple[str, str]:
    text = " ".join(p for _, p in pairs)
    script = "\n".join(f"{r}|{p}" for r, p in pairs)
    return text, script


PREP = {
    1: {
        "name": "manteau",
        "pack": S(
            [
                ("narrateur", "Amir prend le bateau en papier sur la commode."),
                ("narrateur", "Il le glisse au fond du sac. Toc."),
                ("narrateur", "La voile dépasse. Il la rentre du doigt."),
                ("narrateur", "Le manteau est sur la chaise. Il est tiède."),
                ("maman", "Le jardin est encore mouillé."),
                ("maman", "Le manteau, Amir."),
                ("enfant-m", "Oui. Pour le voyage."),
            ]
        ),
        "q": ("Où Amir met-il le bateau ?", "dans le sac", "sac|le sac|dans le sac|au fond du sac", "Le bateau va dans le sac."),
        "c": S(
            [
                ("narrateur", "Le bateau est dans le sac. La voile est rentrée."),
                ("narrateur", "Le manteau attend à côté."),
                ("papa", "On y va, capitaine ?"),
                ("enfant-m", "Oui. Le bateau va choisir."),
            ]
        ),
        "wear_gout": "Amir enfile le manteau. Il est tiède contre ses bras.",
        "wear_choux": "Amir enfile le manteau. Une odeur de soupe reste dans le col.",
        "wear_bac": "Amir enfile le manteau. Le sac tape contre le tissu.",
        "home": "Il raccroche le manteau. Le sac se ferme. Zzz.",
    },
    2: {
        "name": "bottes",
        "pack": S(
            [
                ("narrateur", "Amir glisse le bateau au fond du sac. Toc."),
                ("narrateur", "La voile dépasse. Il la rentre."),
                ("narrateur", "Les bottes sont près de la porte. Un peu mouillées dehors."),
                ("maman", "Attends. Je les essuie."),
                ("narrateur", "Maman passe le linge sur chaque botte."),
                ("maman", "Voilà. Tes pieds resteront au sec."),
            ]
        ),
        "q": ("Qui essuie les bottes ?", "maman", "maman|ma maman|maman essuie", "Maman essuie les bottes."),
        "c": S(
            [
                ("narrateur", "Les bottes sont sèches. Elles attendent près de la porte."),
                ("narrateur", "Le bateau est dans le sac."),
                ("papa", "Capitaine, tes bottes sont prêtes."),
                ("enfant-m", "On va au jardin."),
            ]
        ),
        "wear_gout": "Amir enfile les bottes. Elles sont un peu froides. Clap.",
        "wear_choux": "Amir enfile les bottes. Elles font clap-clap vers les choux.",
        "wear_bac": "Amir enfile les bottes. Elles s'arrêtent pile au bord du bac.",
        "home": "À la porte, il retire les bottes. Le sac se ferme. Zzz.",
    },
    3: {
        "name": "linge",
        "pack": S(
            [
                ("narrateur", "Amir met le bateau dans le sac. Toc."),
                ("narrateur", "La voile rentre sous son doigt."),
                ("narrateur", "Le linge est sur le radiateur. Il sent le propre."),
                ("maman", "Pour tes mains, après l'eau."),
                ("enfant-m", "Je le prends."),
                ("narrateur", "Le linge rejoint le bateau dans le sac."),
            ]
        ),
        "q": ("On met le linge où ?", "dans le sac", "sac|le sac|dans le sac", "Le linge va dans le sac."),
        "c": S(
            [
                ("narrateur", "Le linge est dans le sac. Le bateau aussi."),
                ("papa", "On ouvre la porte ?"),
                ("enfant-m", "Oui. Le voyage commence."),
            ]
        ),
        "wear_gout": "Amir porte le sac. Le linge est dedans, au chaud.",
        "wear_choux": "Amir porte le sac. Le linge sent encore le radiateur.",
        "wear_bac": "Amir porte le sac. Le linge attend les mains mouillées.",
        "home": "Il sèche ses doigts avec le linge. Le linge rentre. Le sac se ferme. Zzz.",
    },
}

OPEN = S(
    [
        ("narrateur", "Le volet jaune bouge encore un peu."),
        ("narrateur", "Dans la rue, les pavés brillent."),
        ("narrateur", "Dans la maison, ça sent la soupe."),
        ("narrateur", "Maman est dans la cuisine."),
        ("narrateur", "Papa est à la grande table. Son ordinateur fait un petit vent."),
        ("narrateur", "Sur le tapis, Amir plie un coin de papier."),
        ("enfant-m", "Celle-ci, la coque."),
        ("narrateur", "Il plie encore. Une voile un peu froissée."),
        ("enfant-m", "Et ça, la cheminée."),
        ("narrateur", "C'est un bateau en papier. Assez petit pour ses deux mains."),
        ("narrateur", "Il le fait glisser. Le bateau s'arrête contre la pantoufle de papa."),
        ("papa", "Votre bateau s'est échoué, capitaine Amir."),
        ("enfant-m", "Il lui faut de l'eau."),
        ("narrateur", "Dehors, la pluie s'arrête. Une dernière goutte. Ploc."),
        ("narrateur", "Amir court à la fenêtre."),
        ("narrateur", "L'eau a dessiné un chemin. La gouttière. Le potager. Le bac à sable."),
        ("enfant-m", "Papa ! Mon bateau peut faire un vrai voyage !"),
        ("papa", "Jusqu'où ?"),
        ("enfant-m", "Je ne sais pas encore. C'est le bateau qui choisira."),
        ("narrateur", "Maman arrive avec le petit sac."),
        ("maman", "Le jardin est encore mouillé. De quoi auras-tu besoin ?"),
    ]
)

T1 = S(
    [
        ("narrateur", "Avant de partir, un capitaine prépare ses affaires."),
        ("narrateur", "Le manteau, les bottes, ou un linge."),
        ("maman", "Qu'est-ce que tu prends, Amir ?"),
    ]
)

T2_Q = S(
    [
        ("narrateur", "Dehors, l'eau a fait trois chemins."),
        ("narrateur", "La rivière de la gouttière. Le port des choux. L'île du bac."),
        ("papa", "Où commence le voyage, capitaine ?"),
    ]
)


def arrive_gout(prep: dict) -> tuple[str, str]:
    return S(
        [
            ("narrateur", prep["wear_gout"]),
            ("narrateur", "Sous la gouttière, l'eau tremble. Plic. Ploc."),
            ("narrateur", "Une grosse feuille barre le passage."),
            ("enfant-m", "La rivière est fermée."),
            ("papa", "Le bateau ne peut pas passer."),
            ("narrateur", "Amir s'accroupit. La feuille ne bouge presque pas."),
        ]
    )


def arrive_choux(prep: dict) -> tuple[str, str]:
    return S(
        [
            ("narrateur", prep["wear_choux"]),
            ("narrateur", "Entre deux choux, une flaque ronde. Comme un port."),
            ("narrateur", "Trois petites feuilles flottent devant l'entrée."),
            ("enfant-m", "Les bateaux n'arrivent pas à accoster."),
            ("papa", "Il y a trois feuilles."),
            ("narrateur", "Amir compte. Une. Deux. Trois."),
        ]
    )


def arrive_bac(prep: dict) -> tuple[str, str]:
    return S(
        [
            ("narrateur", prep["wear_bac"]),
            ("narrateur", "Près du bac à sable, l'eau devient de plus en plus petite."),
            ("narrateur", "Elle rentre dans le sol. Le bateau n'aurait plus assez d'eau."),
            ("enfant-m", "Mon chemin disparaît."),
            ("papa", "Le sable boit la flaque."),
            ("narrateur", "Amir serre le bateau. Il avait imaginé cette île."),
        ]
    )


def t3_gout(res: int) -> tuple[tuple[str, str], tuple[str, str], tuple[str, str, str, str]]:
    if res == 1:
        labels = ("attendre les gouttes", "appeler papa", "changer de départ")
        act = S(
            [
                ("narrateur", "Amir s'accroupit près de la feuille."),
                ("enfant-m", "J'attends encore un peu."),
                ("narrateur", "Une goutte. Plic. La feuille bouge à peine."),
                ("narrateur", "Une deuxième. Ploc. La feuille tourne."),
                ("narrateur", "Une troisième. Plouf. La feuille ouvre un passage."),
                ("enfant-m", "La rivière est libre !"),
            ]
        )
        fin = None  # filled with prep later
        return labels, act, ("gout", "wait")
    if res == 2:
        labels = ("attendre les gouttes", "appeler papa", "changer de départ")
        act = S(
            [
                ("enfant-m", "Papa, tu peux bouger la feuille ?"),
                ("papa", "On la pousse ensemble. Tout doux."),
                ("narrateur", "Papa soulève le bord. Amir tient le bateau."),
                ("narrateur", "La feuille glisse. L'eau circule."),
                ("papa", "À toi, capitaine."),
            ]
        )
        return labels, act, ("gout", "papa")
    act = S(
        [
            ("enfant-m", "On part d'ici, plus loin."),
            ("narrateur", "Amir pose le bateau après la feuille, là où l'eau court déjà."),
            ("narrateur", "Un petit courant emmène le papier sous une brindille."),
            ("enfant-m", "Une autre rivière !"),
        ]
    )
    return ("attendre les gouttes", "appeler papa", "changer de départ"), act, ("gout", "autre")


def t3_choux(res: int):
    labels = ("compter les feuilles", "attendre le courant", "demander à maman")
    if res == 1:
        act = S(
            [
                ("narrateur", "Amir compte encore. Une. Deux. Trois."),
                ("enfant-m", "Là. Un espace libre."),
                ("narrateur", "Il glisse le bateau entre la deuxième et la troisième feuille."),
                ("narrateur", "Le papier touche un quai de terre. Toc."),
                ("enfant-m", "Quai Deux."),
            ]
        )
        return labels, act, ("choux", "count")
    if res == 2:
        act = S(
            [
                ("narrateur", "Amir attend. L'eau pousse tout doux."),
                ("narrateur", "Une feuille s'écarte. Puis une autre."),
                ("enfant-m", "Le port s'ouvre."),
                ("narrateur", "Le bateau avance tout seul jusqu'au chou."),
            ]
        )
        return labels, act, ("choux", "wait")
    act = S(
        [
            ("narrateur", "Maman est venue jusqu'au potager."),
            ("enfant-m", "Maman, je veux jouer sans abîmer les choux."),
            ("maman", "Passe entre les feuilles. Pas dessus."),
            ("narrateur", "Amir pose le bateau dans l'espace libre. Les plants ne bougent pas."),
            ("maman", "Ton port est gentil avec le jardin."),
        ]
    )
    return labels, act, ("choux", "maman")


def t3_bac(res: int):
    labels = ("une autre flaque", "sur le sable", "un autre bateau")
    if res == 1:
        act = S(
            [
                ("enfant-m", "Celle-ci s'en va. On prend l'autre."),
                ("narrateur", "À côté, une flaque plus ronde tient encore."),
                ("narrateur", "Amir y pose le bateau. Il avance. Ffff."),
                ("papa", "Nouveau départ."),
            ]
        )
        return labels, act, ("bac", "autre")
    if res == 2:
        act = S(
            [
                ("enfant-m", "Plus d'eau. Alors il marche."),
                ("narrateur", "Amir pousse le bateau sur le sable humide. Une piste."),
                ("narrateur", "Le papier glisse. Grain. Grain."),
                ("papa", "Un bateau de terre, aujourd'hui."),
            ]
        )
        return labels, act, ("bac", "sable")
    act = S(
        [
            ("enfant-m", "Celui-là est trop fragile."),
            ("narrateur", "Amir pose le bateau en papier dans le sac."),
            ("papa", "On en fabrique un autre, plus solide."),
            ("narrateur", "Avec une feuille plus épaisse, ils plient une nouvelle coque."),
            ("enfant-m", "Celui-ci, pour le sable."),
        ]
    )
    return labels, act, ("bac", "nouveau")


def fin(prep: dict, key: tuple[str, str]) -> tuple[str, str]:
    place, how = key
    souvenir = {
        ("gout", "wait"): [
            ("narrateur", "De retour, Amir pose le bateau sur le petit linge."),
            ("papa", "Alors, capitaine, jusqu'où es-tu allé ?"),
            ("enfant-m", "Très loin. J'ai attendu trois gouttes."),
            ("papa", "C'est vraiment loin."),
            ("narrateur", "Une dernière goutte, dehors. Plic."),
            ("narrateur", "Sur le linge, le bateau attend déjà le prochain voyage."),
        ],
        ("gout", "papa"): [
            ("narrateur", "Dans la maison, Amir pose le bateau près de la fenêtre."),
            ("papa", "On a poussé la feuille ensemble."),
            ("enfant-m", "Oui. La rivière a dit oui."),
            ("narrateur", "Papa reprend sa table. Amir regarde le jardin."),
            ("narrateur", "La gouttière fait encore un tout petit plic."),
        ],
        ("gout", "autre"): [
            ("narrateur", "Amir rentre. Le bateau a une brindille collée à la voile."),
            ("enfant-m", "On a trouvé un autre courant."),
            ("papa", "Le bateau a choisi."),
            ("narrateur", "Amir sourit. Le sac attend près de la porte."),
        ],
        ("choux", "count"): [
            ("narrateur", "À table, Amir dit :"),
            ("enfant-m", "Quai Un. Quai Deux. Quai Trois."),
            ("maman", "Ton bateau a visité le port."),
            ("enfant-m", "Sans déranger les choux."),
            ("narrateur", "Le bateau sèche. Les choux, dehors, restent droits."),
        ],
        ("choux", "wait"): [
            ("narrateur", "Amir raconte :"),
            ("enfant-m", "J'ai attendu. Les feuilles ont ouvert le port."),
            ("papa", "Le courant a aidé."),
            ("narrateur", "Le bateau sent encore un peu le chou. Amir rit."),
        ],
        ("choux", "maman"): [
            ("maman", "Merci d'avoir demandé."),
            ("enfant-m", "Le port est dans le jardin. On le garde."),
            ("narrateur", "Amir pose le bateau. Maman pose la soupe."),
            ("narrateur", "Deux voyages : l'eau, et la maison."),
        ],
        ("bac", "autre"): [
            ("enfant-m", "La première flaque est partie. L'autre m'a emmené."),
            ("papa", "Changer de chemin, ce n'est pas perdre."),
            ("narrateur", "Amir hoche la tête. Le bateau sèche près du volet jaune."),
        ],
        ("bac", "sable"): [
            ("enfant-m", "Aujourd'hui, mon bateau a marché."),
            ("papa", "Un capitaine de sable."),
            ("narrateur", "Un grain reste sur la cheminée rouge. Amir le souffle. Ffff."),
        ],
        ("bac", "nouveau"): [
            ("narrateur", "Deux bateaux maintenant. L'ancien dans le sac. Le nouveau sur la table."),
            ("enfant-m", "Celui-là, pour demain."),
            ("papa", "On a bien fait."),
            ("narrateur", "Le papier fragile se repose. L'autre attend le jardin."),
        ],
    }[key]
    extra = [("narrateur", prep["home"])]
    return S(souvenir + extra)


def main() -> None:
    wb = load_workbook(XLSX)
    meta = wb["meta"]
    for row in meta.iter_rows(min_col=1, max_col=2):
        k = row[0].value
        if k == "title":
            row[1].value = "Le bateau d'Amir et la rivière du jardin"
        if k == "fil_rouge":
            row[1].value = (
                "Amir veut faire voyager son bateau en papier dans le jardin après la pluie. "
                "Il prépare le sac (manteau, bottes ou linge). Puis il choisit un vrai chemin : "
                "rivière de la gouttière, port des choux, ou île du bac. Chaque chemin a un obstacle "
                "et une fin différente. Il rentre. Le bateau attend le prochain voyage."
            )
        if k == "setting":
            row[1].value = "dans la maison, puis le jardin"
    ws = wb["chunks"]
    headers = [c.value for c in ws[1]]
    idx = {n: i + 1 for i, n in enumerate(headers)}

    def put(cid: str, *, text=None, script=None, labels=None, q=None):
        for r in range(2, ws.max_row + 1):
            if ws.cell(r, idx["chunk_id"]).value != cid:
                continue
            if text is not None:
                ws.cell(r, idx["text"], text)
                if "text_ssml" in idx:
                    ws.cell(r, idx["text_ssml"], text)
                if "text_xai_tags" in idx:
                    ws.cell(r, idx["text_xai_tags"], text)
            if script is not None and "script" in idx:
                ws.cell(r, idx["script"], script)
            if labels and "option_1_label" in idx:
                ws.cell(r, idx["option_1_label"], labels[0])
                ws.cell(r, idx["option_2_label"], labels[1])
                ws.cell(r, idx["option_3_label"], labels[2])
            if q and "expected_answer" in idx:
                prompt, ans, acc, retry = q
                if "text" in idx and prompt:
                    pass
                ws.cell(r, idx["expected_answer"], ans)
                ws.cell(r, idx["accepted_examples"], acc)
                if "retry_prompt" in idx:
                    ws.cell(r, idx["retry_prompt"], retry)
            return
        raise SystemExit(f"manque {cid}")

    put("CHK_T0000_P0000", text=OPEN[0], script=OPEN[1])
    put("CHK_T0001_P0000", text=T1[0], script=T1[1], labels=("le manteau", "les bottes", "un linge"))

    t3_fn = {1: t3_gout, 2: t3_choux, 3: t3_bac}
    arrive_fn = {1: arrive_gout, 2: arrive_choux, 3: arrive_bac}

    for p in (1, 2, 3):
        prep = PREP[p]
        base = f"CHK_T0001_P000{p}"
        put(base, text=prep["pack"][0], script=prep["pack"][1])
        put(f"{base}_Q0001", text=prep["q"][0], q=prep["q"])
        put(f"{base}_C0001", text=prep["c"][0], script=prep["c"][1])
        put(f"{base}_T0002_P0000", text=T2_Q[0], script=T2_Q[1], labels=("la rivière de la gouttière", "le port des choux", "l'île du bac"))
        for place in (1, 2, 3):
            arr = arrive_fn[place](prep)
            put(f"{base}_T0002_P000{place}", text=arr[0], script=arr[1])
            labels, _, _ = t3_fn[place](1)
            put(
                f"{base}_T0002_P000{place}_T0003_P0000",
                text=S(
                    [
                        ("narrateur", "Amir regarde."),
                        ("papa", "Comment on fait, capitaine ?"),
                    ]
                )[0],
                script=S(
                    [
                        ("narrateur", "Amir regarde."),
                        ("papa", "Comment on fait, capitaine ?"),
                    ]
                )[1],
                labels=labels,
            )
            for res in (1, 2, 3):
                _lab, act, key = t3_fn[place](res)
                put(f"{base}_T0002_P000{place}_T0003_P000{res}", text=act[0], script=act[1])
                ftxt, fsc = fin(prep, key)
                put(f"{base}_T0002_P000{place}_T0003_P000{res}_F0001", text=ftxt, script=fsc)

    if "journal" in wb.sheetnames:
        wb["journal"].append(["avis1.txt", "fins distinctes gouttière / choux / bac ; ouverture monde ; pas d'audio"])
    wb.save(XLSX)
    wb.close()
    print("ok", XLSX)


if __name__ == "__main__":
    main()
