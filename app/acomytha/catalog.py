"""Import des xlsx atelier vers SQLite."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import delete, func, or_, select
from sqlalchemy.orm import Session

from acomytha.models import Chunk, Lesson, Story
from acomytha.settings import Settings


class CatalogImporter:
    """Lit le référentiel leçons + chaque arbre Excel."""

    CHUNK_FIELDS = (
        "chunk_id",
        "kind",
        "lesson_id",
        "text",
        "option_1_label",
        "option_1_next_chunk",
        "option_2_label",
        "option_2_next_chunk",
        "option_3_label",
        "option_3_next_chunk",
        "default_next_chunk",
        "wait_ms",
        "night_policy",
    )

    def __init__(self, settings: Settings) -> None:
        self._settings = settings

    @property
    def settings(self) -> Settings:
        return self._settings

    def import_all(self, db: Session, limit: int | None = None) -> dict[str, int]:
        n_lessons = self.import_lessons(db)
        n_stories, n_chunks = self.import_stories(db, limit=limit)
        db.commit()
        return {"lessons": n_lessons, "stories": n_stories, "chunks": n_chunks}

    def import_lessons(self, db: Session) -> int:
        path = self.settings.lecons_xlsx
        if not path.exists():
            return 0
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb["lecons"]
        rows = ws.iter_rows(values_only=True)
        header = [str(c) if c else "" for c in next(rows)]
        idx = {name: i for i, name in enumerate(header)}
        count = 0
        for row in rows:
            lid = _cell(row, idx, "lesson_id")
            if not lid:
                continue
            lesson = db.get(Lesson, lid) or Lesson(lesson_id=lid)
            lesson.title = _cell(row, idx, "title")
            lesson.domain_id = _cell(row, idx, "domain_id")
            lesson.domain = _cell(row, idx, "domain")
            lesson.subdomain_id = _cell(row, idx, "subdomain_id")
            lesson.subdomain = _cell(row, idx, "subdomain")
            lesson.framing = _cell(row, idx, "framing") or "standard"
            lesson.objective = _cell(row, idx, "objective")
            db.merge(lesson)
            count += 1
        wb.close()
        db.flush()
        return count

    def import_stories(self, db: Session, limit: int | None = None) -> tuple[int, int]:
        files = sorted(self.settings.arbres_dir.glob("*.xlsx"))
        if limit is not None:
            files = files[:limit]
        n_stories = 0
        n_chunks = 0
        keep: set[str] = set()
        for path in files:
            n_stories += 1
            n_chunks += self._import_one_story(db, path)
            keep.add(path.stem)
        if limit is None:
            self._drop_missing_stories(db, keep)
        return n_stories, n_chunks

    def _drop_missing_stories(self, db: Session, keep: set[str]) -> int:
        """Le live ne sert que les xlsx d'arbres/ (ATOM + TREE-AUT-001). Archive hors catalogue."""
        from acomytha.models import ForestEntry, Purchase

        extra = [s.story_id for s in db.scalars(select(Story)).all() if s.story_id not in keep]
        if not extra:
            return 0
        db.execute(delete(Chunk).where(Chunk.story_id.in_(extra)))
        db.execute(delete(ForestEntry).where(ForestEntry.story_id.in_(extra)))
        db.execute(delete(Purchase).where(Purchase.item_type == "story", Purchase.item_id.in_(extra)))
        db.execute(delete(Story).where(Story.story_id.in_(extra)))
        return len(extra)

    def _import_one_story(self, db: Session, path: Path) -> int:
        wb = load_workbook(path, read_only=True, data_only=True)
        meta = _meta_map(wb["meta"])
        story_id = meta.get("story_id") or path.stem
        audio_dir = self.settings.audio_dir / story_id
        story = db.get(Story, story_id) or Story(story_id=story_id)
        story.editorial_id = meta.get("editorial_id") or story_id
        story.title = meta.get("title") or story_id
        story.kind = meta.get("kind") or "atomic"
        story.age_band = meta.get("age_band") or "N1"
        story.age_range = meta.get("age_range") or ""
        story.lesson_id = meta.get("lesson_id") or ""
        story.secondary_lessons = meta.get("secondary_lessons") or ""
        story.domain = meta.get("domain") or ""
        story.subdomain = meta.get("subdomain") or ""
        story.framing = meta.get("framing") or "standard"
        story.setting = meta.get("setting") or ""
        story.characters = meta.get("characters") or ""
        story.main_character = meta.get("main_character") or _first_character(story.characters)
        story.places = meta.get("places") or _canonical_places(story.setting)
        story.universe = (meta.get("universe") or "").lower()
        story.wait_default_ms = int(meta.get("wait_default_ms") or 3000)
        story.has_audio = (audio_dir / "CHK_T0000_P0000.mp3").exists()
        story.status = "APPROVED_AUDIO" if story.has_audio else "APPROVED_TEXT"
        db.merge(story)
        db.flush()

        db.execute(delete(Chunk).where(Chunk.story_id == story_id))
        ws = wb["chunks"]
        rows = ws.iter_rows(values_only=True)
        header = [str(c) if c else "" for c in next(rows)]
        idx = {name: i for i, name in enumerate(header)}
        n = 0
        for row in rows:
            cid = _cell(row, idx, "chunk_id")
            if not cid:
                continue
            db.add(
                Chunk(
                    chunk_id=cid,
                    story_id=story_id,
                    kind=_cell(row, idx, "kind") or "passage",
                    lesson_id=_cell(row, idx, "lesson_id"),
                    text=_cell(row, idx, "text"),
                    option_1_label=_cell(row, idx, "option_1_label"),
                    option_1_next=_cell(row, idx, "option_1_next_chunk"),
                    option_2_label=_cell(row, idx, "option_2_label"),
                    option_2_next=_cell(row, idx, "option_2_next_chunk"),
                    option_3_label=_cell(row, idx, "option_3_label"),
                    option_3_next=_cell(row, idx, "option_3_next_chunk"),
                    default_next=_cell(row, idx, "default_next_chunk"),
                    wait_ms=int(row[idx["wait_ms"]] or 0) if "wait_ms" in idx else 0,
                    night_policy=_cell(row, idx, "night_policy") or "play",
                )
            )
            n += 1
        story.chunk_count = n
        db.merge(story)
        wb.close()
        return n


def story_to_dict(story: Story) -> dict:
    return {
        "story_id": story.story_id,
        "title": story.title,
        "kind": story.kind,
        "age_band": story.age_band,
        "age_range": story.age_range,
        "lesson_id": story.lesson_id,
        "secondary_lessons": [s for s in (story.secondary_lessons or "").split("|") if s.strip()],
        "domain": story.domain,
        "subdomain": story.subdomain,
        "framing": story.framing,
        "setting": story.setting,
        "characters": story.characters,
        "main_character": story.main_character,
        "places": [value.strip() for value in (story.places or "").split("|") if value.strip()],
        "universe": story.universe,
        "chunk_count": story.chunk_count,
        "duration_s": story.duration_s or 0,
        "has_interaction": bool(story.has_interaction),
        "has_audio": story.has_audio,
        "status": story.status,
        "wait_default_ms": story.wait_default_ms,
    }


def fill_durations(db: Session, settings: Settings, force: bool = False) -> int:
    """Estime la durée d’écoute (chemin typique) à partir du texte, une fois."""
    from collections import defaultdict

    need = [s for s in db.scalars(select(Story)).all() if force or not s.duration_s]
    if not need:
        return 0
    words: dict[str, list[int]] = defaultdict(lambda: [0, 0])
    for sid, text in db.query(Chunk.story_id, Chunk.text):
        words[sid][0] += len((text or "").split())
        words[sid][1] += 1
    n = 0
    for story in need:
        w, c = words.get(story.story_id, [40, 5])
        if story.kind == "ramifiee":
            w = max(40, w // 9)
            c = max(8, c // 9)
        audio_dir = settings.audio_dir / story.story_id
        mp3s = list(audio_dir.glob("*.mp3")) if audio_dir.exists() else []
        if mp3s and story.kind != "ramifiee":
            sec = sum(max(1, p.stat().st_size * 8 // 64000) for p in mp3s)
        else:
            sec = int(w / 2.0 + c * 0.8)
        story.duration_s = max(45, min(int(sec), 720))
        n += 1
    db.commit()
    return n


def fill_interaction(db: Session) -> int:
    """Marque les histoires qui ont un passage-question ou un choix."""
    ids = {
        sid
        for (sid,) in db.query(Chunk.story_id)
        .filter(Chunk.kind.ilike("%question%") | Chunk.kind.ilike("%choice%"))
        .distinct()
    }
    n = 0
    for story in db.scalars(select(Story)).all():
        flag = story.story_id in ids or story.kind == "ramifiee"
        if story.has_interaction != flag:
            story.has_interaction = flag
            n += 1
    db.commit()
    return n


def list_stories(db: Session, q: str = "", domain: str = "", age_band: str = "", kind: str = "") -> list[Story]:
    stmt = _story_filter(q=q, domain=domain, age_band=age_band, kind=kind)
    stmt = stmt.order_by(Story.domain, Story.age_band, Story.title)
    return list(db.scalars(stmt))


def page_stories(
    db: Session,
    q: str = "",
    domain: str = "",
    age_band: str = "",
    kind: str = "",
    characters: str = "",
    lessons: str = "",
    places: str = "",
    universes: str = "",
    limit: int = 6,
    offset: int = 0,
) -> tuple[list[Story], int]:
    limit = max(1, min(int(limit), 48))
    offset = max(0, int(offset))
    filtered = _story_filter(q=q, domain=domain, age_band=age_band, kind=kind, characters=characters, lessons=lessons, places=places, universes=universes)
    total = int(db.scalar(select(func.count()).select_from(filtered.subquery())) or 0)
    stmt = filtered.order_by(Story.domain, Story.age_band, Story.title).offset(offset).limit(limit)
    return list(db.scalars(stmt)), total


def related_for(db: Session, stories: list[Story]) -> dict[str, list[dict]]:
    lids = {s.lesson_id for s in stories if s.kind == "ramifiee" and s.lesson_id}
    if not lids:
        return {}
    rows = list(db.scalars(select(Story).where(Story.lesson_id.in_(lids))))
    by: dict[str, list[dict]] = {}
    for r in rows:
        by.setdefault(r.lesson_id, []).append({"story_id": r.story_id, "title": r.title})
    return by


def _story_filter(q: str = "", domain: str = "", age_band: str = "", kind: str = "", characters: str = "", lessons: str = "", places: str = "", universes: str = ""):
    stmt = select(Story)
    if q:
        like = f"%{q.strip()}%"
        stmt = stmt.where(
            or_(
                Story.title.ilike(like),
                Story.story_id.ilike(like),
                Story.lesson_id.ilike(like),
                Story.setting.ilike(like),
                Story.characters.ilike(like),
            )
        )
    if domain:
        stmt = stmt.where(Story.domain == domain.upper())
    if age_band:
        stmt = stmt.where(Story.age_band == age_band.upper())
    if kind == "interaction":
        stmt = stmt.where(Story.has_interaction.is_(True), Story.kind != "ramifiee")
    elif kind == "ramifiee":
        stmt = stmt.where(Story.kind == "ramifiee")
    elif kind:
        stmt = stmt.where(Story.kind == kind)
    for raw, column in ((characters, Story.characters), (places, Story.places)):
        values = [value.strip() for value in raw.split(",") if value.strip()]
        if values:
            stmt = stmt.where(or_(*(column.ilike(f"%{value}%") for value in values)))
    lesson_values = [value.strip() for value in lessons.split(",") if value.strip()]
    if lesson_values:
        stmt = stmt.where(Story.lesson_id.in_(lesson_values))
    universe_values = [value.strip().lower() for value in universes.split(",") if value.strip()]
    if universe_values:
        stmt = stmt.where(Story.universe.in_(universe_values))
    return stmt


def _first_character(value: str) -> str:
    return value.replace("|", ",").split(",", 1)[0].strip()


def _canonical_places(setting: str) -> str:
    folded = setting.casefold()
    aliases = {
        "maison": ("maison", "salon", "cuisine", "chambre", "salle de bain"),
        "école": ("école", "classe", "cour de récréation", "cantine"),
        "parc": ("parc", "square", "aire de jeux"),
        "jardin": ("jardin", "potager"),
        "bibliothèque": ("bibliothèque", "médiathèque"),
        "plage": ("plage", "bord de mer"),
        "commerces": ("marché", "magasin", "boulangerie", "supermarché"),
        "transports": ("gare", "train", "bus", "métro", "tramway"),
    }
    return " | ".join(label for label, words in aliases.items() if any(word in folded for word in words))


def _meta_map(ws) -> dict[str, str]:
    out: dict[str, str] = {}
    for row in ws.iter_rows(values_only=True):
        if not row or row[0] in (None, "clé"):
            continue
        out[str(row[0])] = "" if row[1] is None else str(row[1])
    return out


def _cell(row, idx: dict[str, int], name: str) -> str:
    if name not in idx:
        return ""
    val = row[idx[name]]
    if val is None:
        return ""
    return str(val).strip()
