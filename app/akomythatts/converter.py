"""Job Excel → JSON schema 2.0, compatible TTS et empreintes vocales."""

from __future__ import annotations

import hashlib
import json
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .catalogue import CharacterCatalogue, TROUPE
from .detector import CharacterDetector
from .settings import Settings
from .utils import Utils

SCHEMA_VERSION = "2.0"
ROLE_ALIASES = Utils.ROLE_ALIASES
ENERGY_TO_EMOTION = {
    "calm": "calm",
    "warm": "warm",
    "bright": "joy",
    "focused": "focused",
    "tense": "suspense",
    "lively": "excited",
    "storytelling": "storytelling",
}
CONTOUR_TO_INTONATION = {
    "level": "neutral",
    "fall": "falling",
    "falling": "falling",
    "rise": "rising",
    "rising": "rising",
    "dynamic": "dramatic",
    "storytelling": "storytelling",
}
_PART = re.compile(r"^([A-Z]+)(\d+)?$")
_RANK_HEAD = {"T": 0, "P": 1, "O": 2, "Q": 3, "C": 4, "END": 6, "F": 7}


class CatalogueConverter:
    """Lit un xlsx AcoMytha et écrit un JSON TTS (chunks + speaker_profiles)."""

    def __init__(self, settings: Settings, catalogue: CharacterCatalogue) -> None:
        self.settings = settings
        self.catalogue = catalogue
        self.detector = CharacterDetector(catalogue)

    def convert_directory(self, source: Path, output: Path | None = None) -> dict[str, Any]:
        output = output or (self.settings.root / "stories" / "json")
        output.mkdir(parents=True, exist_ok=True)
        files = sorted(source.glob("*.xlsx"))
        report = {"ok": 0, "errors": [], "files": []}
        for path in files:
            try:
                story = self.convert_file(path)
                dest = output / f"{story['story_id']}.json"
                dest.write_text(json.dumps(story, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
                report["ok"] += 1
                report["files"].append(
                    {
                        "story_id": story["story_id"],
                        "title": story["title"],
                        "chunks": len(story["chunks"]),
                        "speakers": story["speaker_profiles"],
                        "json": str(dest.relative_to(self.settings.root)),
                    }
                )
            except Exception as exc:
                report["errors"].append(f"{path.name}: {exc}")
        self.catalogue.save()
        (output / "conversion_report.json").write_text(
            json.dumps(report, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
        )
        return report

    def convert_file(self, path: Path) -> dict[str, Any]:
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            meta = {
                self._clean(row[0].value): self._clean(row[1].value)
                for row in workbook["meta"].iter_rows(min_row=2)
                if row[0].value
            }
            story_id = meta.get("story_id") or path.stem
            hero = self._hero(meta.get("characters") or "")
            rows = list(self._rows(workbook))
            chunks: dict[str, Any] = {}
            for row in rows:
                chunk_id = self._clean(row.get("chunk_id"))
                if not chunk_id:
                    continue
                options = []
                for number in range(1, 4):
                    label = self._clean(row.get(f"option_{number}_label"))
                    target = self._clean(row.get(f"option_{number}_next_chunk"))
                    if label or target:
                        options.append({"id": str(number), "label": label, "next_chunk": target})
                chunks[chunk_id] = {
                    "kind": self._clean(row.get("kind")) or "passage",
                    "lesson_id": self._clean(row.get("lesson_id")),
                    "segments": self._parse_script(row, hero, story_id),
                    "interaction": {
                        "expected_answer": self._clean(row.get("expected_answer")),
                        "accepted_examples": [
                            item.strip()
                            for item in self._clean(row.get("accepted_examples")).split("|")
                            if item.strip()
                        ],
                        "correct_response": self._clean(row.get("engine_ok_text")),
                        "near_response": self._clean(row.get("engine_near_text")),
                        "retry_prompt": self._clean(row.get("retry_prompt")),
                        "retry_once": self._clean(row.get("retry_once")).lower() == "oui",
                        "wait_ms": int(row.get("wait_ms") or 0),
                    },
                    "options": options,
                    "source_default_next_chunk": self._clean(row.get("default_next_chunk")) or None,
                    "night_policy": self._clean(row.get("night_policy")) or "play",
                    "sound_cues": [
                        item.strip() for item in self._clean(row.get("sons")).replace(",", "|").split("|") if item.strip()
                    ],
                    "editorial_notes": self._clean(row.get("notes")),
                }
            linear = all(not value["options"] and not str(value["kind"]).startswith("transition") for value in chunks.values())
            for chunk_id, chunk in chunks.items():
                chunk["next_chunk"] = self._successor(chunk_id, chunks, linear)
                default = chunk.pop("source_default_next_chunk")
                chunk["default_next_chunk"] = default or chunk["next_chunk"]
            speakers = sorted({seg["speaker"] for chunk in chunks.values() for seg in chunk["segments"]})
            return {
                "schema_version": SCHEMA_VERSION,
                "story_id": story_id,
                "title": meta.get("title") or story_id,
                "language": meta.get("language") or "fr",
                "catalogue": {
                    "kind": meta.get("kind") or "atomic",
                    "age_band": meta.get("age_band"),
                    "age_range": meta.get("age_range"),
                    "lesson_id": meta.get("lesson_id"),
                    "secondary_lessons": [
                        item.strip() for item in meta.get("secondary_lessons", "").split("|") if item.strip()
                    ],
                    "domain": meta.get("domain"),
                    "subdomain": meta.get("subdomain"),
                    "setting": meta.get("setting"),
                    "characters": [item.strip() for item in meta.get("characters", "").split(",") if item.strip()],
                },
                "source": {
                    "xlsx": str(path.relative_to(self.settings.root))
                    if path.is_relative_to(self.settings.root)
                    else str(path),
                    "sha256": hashlib.sha256(path.read_bytes()).hexdigest(),
                },
                "editorial_status": "chatgpt_corrected_ready_for_tts",
                "entry_chunk": next(iter(chunks), None),
                "speaker_profiles": speakers,
                "chunks": chunks,
            }
        finally:
            workbook.close()

    def _parse_script(self, row: dict[str, Any], hero: str, story_id: str) -> list[dict[str, Any]]:
        script = self._clean(row.get("script"))
        lines = script.splitlines() if script else [f"narrateur|{self._clean(row.get('text'))}"]
        segments = []
        for line in lines:
            if not line.strip():
                continue
            role, separator, text = line.partition("|")
            if not separator:
                role, text = "narrateur", line
            text = text.strip()
            if not text:
                continue
            speaker = self._speaker_id(role, hero, story_id)
            segments.append(
                {
                    "speaker": speaker,
                    "text": text,
                    "prosody": self._prosody(row, text, role.strip().lower()),
                }
            )
        return segments

    def _speaker_id(self, raw_role: str, hero: str, story_id: str) -> str:
        role = raw_role.strip().lower()
        if role in {"enfant-m", "enfant-f"}:
            return f"character.{self.slug(hero)}"
        if role in ROLE_ALIASES:
            return ROLE_ALIASES[role]
        slug = self.slug(raw_role)
        if slug in TROUPE:
            return f"character.{slug}"
        return f"story.{self.slug(story_id)}.{slug}"

    def _prosody(self, row: dict[str, Any], text: str, role: str) -> dict[str, Any]:
        speed = float(row.get("kokoro_speed") or row.get("speed_xai") or 0.94)
        gain = float(row.get("volume_db") or 0.0)
        pause = int(row.get("pause_sentence_ms") or row.get("pause_after_ms") or 280)
        emotion = ENERGY_TO_EMOTION.get(self._clean(row.get("style_energy")).lower(), "neutral")
        intonation = CONTOUR_TO_INTONATION.get(self._clean(row.get("style_contour")).lower(), "neutral")
        if "?" in text:
            intonation = "rising"
            pause = max(pause, 340)
        if "!" in text:
            emotion = "excited" if role.startswith("enfant") else "joy"
            speed = min(speed + 0.04, 1.2)
        kind = self._clean(row.get("kind"))
        if kind == "passage_debut" and role in {"narrateur", "narratrice"}:
            emotion, intonation = "storytelling", "storytelling"
        elif kind == "passage_fin":
            emotion, intonation = "warm", "falling"
            pause = max(pause, 500)
        elif "question" in kind:
            intonation = "rising"
        words = [item.strip() for item in self._clean(row.get("emphasis_words")).replace(",", "|").split("|") if item.strip()]
        words = [word for word in words if word.casefold() in text.casefold()]
        return {
            "speed": round(max(0.65, min(speed, 1.25)), 3),
            "gain_db": round(max(-6.0, min(gain, 5.0)), 2),
            "pitch_semitones": 0.0,
            "pause_before_ms": int(row.get("pause_before_ms") or 0),
            "pause_after_ms": pause,
            "emotion": emotion,
            "intonation": intonation,
            "emphasis_words": words,
        }

    def _successor(self, chunk_id: str, chunks: dict[str, Any], linear: bool) -> str | None:
        chunk = chunks[chunk_id]
        if chunk["kind"] == "passage_fin":
            return None
        explicit = chunk.get("source_default_next_chunk")
        if explicit in chunks:
            return explicit
        if linear:
            ordered = sorted(chunks, key=self._chunk_sort_key)
            index = ordered.index(chunk_id)
            return ordered[index + 1] if index + 1 < len(ordered) else None
        if chunk_id == "CHK_T0000_P0000" and "CHK_T0001_P0000" in chunks:
            return "CHK_T0001_P0000"
        return None

    def _chunk_sort_key(self, chunk_id: str) -> tuple[Any, ...]:
        body = chunk_id[4:] if chunk_id.startswith("CHK_") else chunk_id
        key = []
        for index, part in enumerate(body.split("_")):
            match = _PART.match(part)
            if not match:
                key.append((50, 0))
                continue
            letters, number = match.group(1), int(match.group(2) or 0)
            rank = 5 if letters == "T" and index > 0 else _RANK_HEAD.get(letters, 40)
            key.append((rank, number))
        return tuple(key)

    def _rows(self, workbook: Any) -> list[dict[str, Any]]:
        rows = workbook["chunks"].iter_rows(values_only=True)
        headers = [self._clean(value) for value in next(rows)]
        out = []
        for values in rows:
            row = dict(zip(headers, values))
            if self._clean(row.get("chunk_id")):
                out.append(row)
        return out

    def _hero(self, raw: str) -> str:
        parts = [part.strip() for part in re.split(r"[,;|/]| et ", raw) if part.strip()]
        skip = {"papa", "maman", "père", "mère", "narrateur", "narratrice"}
        named = [part for part in parts if part.casefold() not in skip]
        for part in named:
            slug = re.sub(r"[^a-z0-9]+", "", self.slug(part).replace("_", ""))
            if slug in TROUPE:
                return TROUPE[slug][0]
        return named[0] if named else "Enfant"

    @staticmethod
    def _clean(value: Any) -> str:
        return "" if value is None else str(value).strip()

    @staticmethod
    def slug(value: str) -> str:
        return Utils.file_slug(value)
