"""Parse un JSON ou un Excel d'histoire en objet ParsedStory."""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from .models import ParsedStory
from .utils import Utils


class StoryParser:
    """Lit un JSON schema 2.0, un JSON simple, ou un Excel atelier."""

    def parse_bytes(self, filename: str, payload: bytes) -> ParsedStory:
        name = filename.lower()
        if name.endswith(".xlsx"):
            return self.parse_xlsx(payload)
        text = payload.decode("utf-8")
        data = json.loads(text)
        if not isinstance(data, dict):
            raise ValueError("Le JSON doit être un objet.")
        return self.parse_json(data, Path(filename).stem)

    def parse_path(self, path: Path) -> ParsedStory:
        return self.parse_bytes(path.name, path.read_bytes())

    def parse_json(self, data: dict[str, Any], fallback_title: str = "histoire") -> ParsedStory:
        if data.get("segments"):
            speakers = dict(data.get("speakers") or {})
            segments = list(data["segments"])
            for raw in segments:
                speakers.setdefault(str(raw.get("speaker") or "narrator"), {})
            excerpt = str(segments[0].get("text") or "")[:180] if segments else ""
            return ParsedStory(
                title=str(data.get("title") or fallback_title),
                format="simple",
                speaker_keys=list(speakers),
                segments=segments,
                characters_hint=str(data.get("characters") or ""),
                excerpt=excerpt,
            )
        chunks = data.get("chunks")
        if not isinstance(chunks, dict) or not chunks:
            raise ValueError("JSON invalide : ni segments ni chunks.")
        speaker_keys: list[str] = []
        for profile in data.get("speaker_profiles") or []:
            if str(profile) not in speaker_keys:
                speaker_keys.append(str(profile))
        for name in data.get("speakers") or {}:
            if name not in speaker_keys:
                speaker_keys.append(name)
        chunk_id = data.get("entry_chunk") or next(iter(chunks))
        seen: set[str] = set()
        segments: list[dict[str, Any]] = []
        used = 0
        while chunk_id and chunk_id not in seen:
            seen.add(chunk_id)
            chunk = chunks.get(chunk_id)
            if not isinstance(chunk, dict):
                break
            used += 1
            for raw in chunk.get("segments") or []:
                key = str(raw.get("speaker") or "narrator")
                if key not in speaker_keys:
                    speaker_keys.append(key)
                segments.append(raw)
            nxt = chunk.get("default_next_chunk") or chunk.get("next_chunk")
            if isinstance(nxt, list):
                chunk_id = str(nxt[0]) if nxt else None
            elif isinstance(nxt, str) and nxt:
                chunk_id = nxt
            else:
                chunk_id = None
        if not segments:
            raise ValueError("Aucun segment dans le parcours.")
        catalogue = data.get("catalogue") if isinstance(data.get("catalogue"), dict) else {}
        names = catalogue.get("characters") or data.get("characters") or []
        if isinstance(names, list):
            hint = ", ".join(str(name) for name in names if str(name).strip())
        else:
            hint = str(names or "").strip()
        return ParsedStory(
            title=str(data.get("title") or data.get("story_id") or fallback_title),
            format="catalogue",
            speaker_keys=speaker_keys,
            segments=segments,
            characters_hint=hint or ", ".join(speaker_keys),
            excerpt=str(segments[0].get("text") or "")[:180],
            chunks=used,
        )

    def parse_xlsx(self, payload: bytes) -> ParsedStory:
        from io import BytesIO

        import openpyxl

        wb = openpyxl.load_workbook(BytesIO(payload), read_only=True, data_only=True)
        meta = {row[0]: row[1] for row in wb["meta"].iter_rows(values_only=True) if row and row[0]}
        header = [cell for cell in next(wb["chunks"].iter_rows(values_only=True))]
        idx = {name: i for i, name in enumerate(header) if name}
        segments: list[dict[str, Any]] = []
        speaker_keys: list[str] = []
        for row in wb["chunks"].iter_rows(values_only=True):
            if not row or row[0] == "chunk_id":
                continue
            script = row[idx["script"]] if "script" in idx else None
            text = row[idx["text"]] if "text" in idx else None
            if script:
                for line in str(script).splitlines():
                    if "|" not in line:
                        continue
                    role, replica = line.split("|", 1)
                    role = role.strip()
                    replica = replica.strip()
                    if not replica:
                        continue
                    key = Utils.speaker_key(role)
                    if key not in speaker_keys:
                        speaker_keys.append(key)
                    segments.append({"speaker": key, "text": replica, "prosody": self._prosody_from_row(row, idx)})
            elif text:
                key = "narrator"
                if key not in speaker_keys:
                    speaker_keys.append(key)
                segments.append({"speaker": key, "text": str(text).strip(), "prosody": self._prosody_from_row(row, idx)})
        wb.close()
        if not segments:
            raise ValueError("Excel sans répliques lisibles.")
        return ParsedStory(
            title=str(meta.get("title") or meta.get("story_id") or "Histoire"),
            format="xlsx",
            speaker_keys=speaker_keys,
            segments=segments,
            characters_hint=str(meta.get("characters") or ""),
            excerpt=segments[0]["text"][:180],
            chunks=int(meta.get("chunk_count") or 0),
        )

    def _role_to_key(self, role: str) -> str:
        return Utils.speaker_key(role)

    def _prosody_from_row(self, row: tuple[Any, ...], idx: dict[str, int]) -> dict[str, Any]:
        def val(name: str, default: Any) -> Any:
            if name not in idx:
                return default
            item = row[idx[name]]
            return default if item is None else item

        energy = str(val("style_energy", "neutral") or "neutral")
        emotion_map = {
            "calm": "calm",
            "warm": "warm",
            "lively": "joy",
            "focused": "focused",
            "storytelling": "storytelling",
        }
        contour = str(val("style_contour", "neutral") or "neutral")
        intonation_map = {"rise": "rising", "rising": "rising", "fall": "falling", "dramatic": "dramatic", "storytelling": "storytelling"}
        return {
            "speed": float(val("kokoro_speed", 1.0) or 1.0),
            "gain_db": float(val("volume_db", 0.0) or 0.0),
            "pause_before_ms": int(val("pause_before_ms", 0) or 0),
            "pause_after_ms": int(val("pause_after_ms", 250) or 250),
            "emotion": emotion_map.get(energy, "neutral"),
            "intonation": intonation_map.get(contour, "neutral"),
        }
